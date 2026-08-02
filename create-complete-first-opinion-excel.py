#!/usr/bin/env python3
"""
DISHA First Opinion Engine - Complete Excel Workbook
Shows entire user workflow with data mapping, calculations, and questionnaires
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_color_scheme():
    """Define consistent color scheme"""
    return {
        'header_dark': 'FF1E40AF',        # Dark blue
        'header_light': 'FF0D9488',        # Teal
        'input_field': 'FFCCFBF1',         # Light cyan
        'calculation': 'FFFEF3C7',         # Yellow
        'result': 'FFFFE4E6',              # Light red/pink
        'success': 'FFDBEAFE',             # Light green
        'text_white': 'FFFFFFFF',
        'text_dark': 'FF1E293B',
        'divider': 'FFE2E8F0'
    }

def apply_header_style(cell, colors, color_key='header_dark'):
    """Apply header styling to a cell"""
    cell.font = Font(bold=True, color='FFFFFFFF', size=12)
    cell.fill = PatternFill(start_color=colors[color_key], end_color=colors[color_key], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def apply_subheader_style(cell, colors):
    """Apply subheader styling"""
    cell.font = Font(bold=True, color='FFFFFFFF', size=11)
    cell.fill = PatternFill(start_color=colors['header_light'], end_color=colors['header_light'], fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_input_style(cell, colors):
    """Apply input field styling"""
    cell.fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_calc_style(cell, colors):
    """Apply calculation field styling"""
    cell.fill = PatternFill(start_color=colors['calculation'], end_color=colors['calculation'], fill_type='solid')
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_result_style(cell, colors):
    """Apply result styling"""
    cell.fill = PatternFill(start_color=colors['result'], end_color=colors['result'], fill_type='solid')
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def create_workflow_sheet(wb, colors):
    """Sheet 1: User Workflow Overview"""
    ws = wb.create_sheet('01_Workflow Overview', 0)

    ws['A1'] = 'DISHA FIRST OPINION ENGINE - USER WORKFLOW'
    apply_header_style(ws['A1'], colors)
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 25

    ws['A3'] = 'STEP-BY-STEP USER WORKFLOW'
    apply_subheader_style(ws['A3'], colors)
    ws.merge_cells('A3:F3')

    workflow_steps = [
        ('STEP 1', 'School Registration', 'Enter school name, board, location, student count, fees'),
        ('STEP 2', 'Challenge Selection', 'Select 1-3 primary operational challenges from 15 options'),
        ('STEP 3', 'Dynamic Screening', 'Answer 3 questions per selected challenge (customized per challenge)'),
        ('STEP 4', 'Operational Metrics', 'Enter 4 key metrics: STR, Parent SLA, Training, Planning hours'),
        ('STEP 5', 'Evidence Upload', 'Optional: Upload CSV samples (attendance, fees, staff)'),
        ('STEP 6', 'Auto-Calculation', 'System calculates: S_sub, M_obj, P_mismatch'),
        ('STEP 7', 'Diagnosis & Output', 'Get health index (0-100), risk quadrant, prescriptive actions')
    ]

    row = 4
    for step, title, description in workflow_steps:
        ws[f'A{row}'] = step
        ws[f'A{row}'].font = Font(bold=True, color='FFFFFFFF', size=11)
        ws[f'A{row}'].fill = PatternFill(start_color=colors['header_light'], end_color=colors['header_light'], fill_type='solid')

        ws[f'B{row}'] = title
        ws[f'B{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'].fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type='solid')

        ws[f'C{row}'] = description
        ws[f'C{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'C{row}:F{row}')

        ws.row_dimensions[row].height = 30
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60

def create_challenge_catalog(wb, colors):
    """Sheet 2: All 15 Challenges & Categories"""
    ws = wb.create_sheet('02_15 Challenges', 1)

    ws['A1'] = 'DISHA CHALLENGE CATALOG - 15 INSTITUTIONAL CHALLENGES'
    apply_header_style(ws['A1'], colors)
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 25

    # Headers
    headers = ['Category', 'Challenge ID', 'Challenge Name', 'Primary Metric', 'Data Source', 'Benchmark']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_header_style(cell, colors)

    challenges_data = [
        ('Growth & Enrollment', 'C1', 'Enrollment Decline', 'Inquiry-to-admission %', 'Inquiry logs', '>25%'),
        ('Growth & Enrollment', 'C2', 'Student Attrition', 'Annual dropout %', 'Enrollment records', '<5%'),
        ('Growth & Enrollment', 'C3', 'Fee Collection', 'Fee default %', 'Finance ledger', '<5%'),
        ('People & Staffing', 'C4', 'Teacher Attrition', 'Annual turnover %', 'HR records', '<10%'),
        ('People & Staffing', 'C5', 'Staff Capability', 'Training hours/year', 'CPD logs', '>=25 hrs'),
        ('People & Staffing', 'C6', 'Leadership Gap', 'Pipeline ratio', 'Org chart', '3+ people'),
        ('Academic & Wellbeing', 'C7', 'Academic Decline', 'Board pass %', 'Board results', '>90%'),
        ('Academic & Wellbeing', 'C8', 'Student Wellbeing', 'Wellness index', 'Surveys', 'High'),
        ('Academic & Wellbeing', 'C9', 'Remedial Gap', 'Coverage %', 'Academic records', '>90%'),
        ('Reputation & Competition', 'C10', 'Parent Communication', 'Response time (hrs)', 'Ticketing', '<=12 hrs'),
        ('Reputation & Competition', 'C11', 'Competitor Pressure', 'Market position', 'Market analysis', 'Strong'),
        ('Reputation & Competition', 'C12', 'Brand Perception', 'NPS score', 'Parent surveys', '>50'),
        ('Operations & Finance', 'C13', 'Cost Inflation', 'Cost-to-revenue %', 'Finance', '<65%'),
        ('Operations & Finance', 'C14', 'Infrastructure', 'Facility condition', 'Audit', 'Excellent'),
        ('Operations & Finance', 'C15', 'Compliance', 'Audit score', 'Board records', '100%')
    ]

    row = 4
    for category, cid, name, metric, source, benchmark in challenges_data:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = cid
        ws[f'C{row}'] = name
        ws[f'D{row}'] = metric
        ws[f'E{row}'] = source
        ws[f'F{row}'] = benchmark

        ws[f'A{row}'].fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type='solid')

        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].alignment = Alignment(wrap_text=True)

        ws.row_dimensions[row].height = 25
        row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

def create_questionnaire_master(wb, colors):
    """Sheet 3: Complete Questionnaire for All 15 Challenges"""
    ws = wb.create_sheet('03_Master Questionnaire', 2)

    ws['A1'] = 'COMPLETE SCREENING QUESTIONNAIRE - ALL 15 CHALLENGES'
    apply_header_style(ws['A1'], colors)
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 25

    questionnaire_data = {
        'C1: Enrollment Decline': [
            ('Q1.1', 'Inquiry-to-Admission Conversion', 'Above 25%', 2, 'Healthy conversion'),
            ('Q1.1', 'Inquiry-to-Admission Conversion', '10-25%', 6, 'Suboptimal conversion'),
            ('Q1.1', 'Inquiry-to-Admission Conversion', 'Below 10%', 10, 'Critical deficit'),
            ('Q1.2', 'Primary Marketing Channels', 'Digital (Google, Meta)', 3, 'Modern, trackable'),
            ('Q1.2', 'Primary Marketing Channels', 'Traditional (Print)', 6, 'Limited ROI tracking'),
            ('Q1.2', 'Primary Marketing Channels', 'No formal budget', 8, 'Word-of-mouth only'),
            ('Q1.3', 'Parent Drop-off Stage', 'After initial inquiry', 9, 'Communication issue'),
            ('Q1.3', 'Parent Drop-off Stage', 'After touring school', 7, 'Value perception gap'),
            ('Q1.3', 'Parent Drop-off Stage', 'After final offer', 8, 'Competitive loss'),
        ],
        'C4: Teacher Attrition': [
            ('Q4.1', 'Annual Teacher Turnover Rate', 'Under 10%', 2, 'Stable retention'),
            ('Q4.1', 'Annual Teacher Turnover Rate', '10-25%', 6, 'Moderate churn'),
            ('Q4.1', 'Annual Teacher Turnover Rate', 'Above 25%', 10, 'Severe instability'),
            ('Q4.2', 'Teaching Load (periods/week)', '18-22 periods', 2, 'Sustainable'),
            ('Q4.2', 'Teaching Load (periods/week)', '24-28 periods', 6, 'Heavy load'),
            ('Q4.2', 'Teaching Load (periods/week)', '30+ periods', 9, 'Burnout risk'),
            ('Q4.3', 'Exit Interview Reason', 'Career gap', 5, 'Growth opportunity'),
            ('Q4.3', 'Exit Interview Reason', 'Better salary', 8, 'Compensation gap'),
            ('Q4.3', 'Exit Interview Reason', 'Burnout', 9, 'Systemic issue'),
        ],
        'C10: Parent Communication': [
            ('Q10.1', 'Communication Channel', 'Formal portal/ticketing', 2, 'Structured, trackable'),
            ('Q10.1', 'Communication Channel', 'Physical/diary notes', 6, 'Semi-formal'),
            ('Q10.1', 'Communication Channel', 'WhatsApp/direct calls', 9, 'Chaotic communication'),
            ('Q10.2', 'Query Response Time (SLA)', 'Under 24 hours', 1, 'Rapid response'),
            ('Q10.2', 'Query Response Time (SLA)', '24-48 hours', 5, 'Acceptable delay'),
            ('Q10.2', 'Query Response Time (SLA)', 'Over 48 hours', 9, 'Severe communication breakdown'),
        ]
    }

    row = 3
    for challenge, questions in questionnaire_data.items():
        # Challenge header
        ws[f'A{row}'] = challenge
        apply_subheader_style(ws[f'A{row}'], colors)
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 20
        row += 1

        # Column headers
        col_headers = ['Q ID', 'Question', 'Option', 'Weight', 'Interpretation']
        for col, header in enumerate(col_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color='FFE2E8F0', end_color='FFE2E8F0', fill_type='solid')
        row += 1

        # Questions
        for qid, question, option, weight, interpretation in questions:
            ws[f'A{row}'] = qid
            ws[f'B{row}'] = question
            ws[f'C{row}'] = option
            ws[f'D{row}'] = weight
            ws[f'E{row}'] = interpretation

            ws[f'D{row}'].fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type='solid')
            ws[f'D{row}'].font = Font(bold=True)

            for col in ['B', 'C', 'E']:
                ws[f'{col}{row}'].alignment = Alignment(wrap_text=True)

            ws.row_dimensions[row].height = 20
            row += 1

        row += 1  # Blank row between challenges

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 25

def create_example_scenario(wb, colors):
    """Sheet 4: Example Scenario with Hardcoded Data & Formulas"""
    ws = wb.create_sheet('04_Example Calculation', 3)

    ws['A1'] = 'DISHA FIRST OPINION ENGINE - LIVE CALCULATION EXAMPLE'
    apply_header_style(ws['A1'], colors)
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 25

    # School Details
    row = 3
    ws[f'A{row}'] = 'SCHOOL DETAILS'
    apply_subheader_style(ws[f'A{row}'], colors)
    ws.merge_cells(f'A{row}:D{row}')

    school_data = [
        ('School Name', 'ABC High School'),
        ('Board Affiliation', 'CBSE'),
        ('Total Students', 450),
        ('City/District', 'Mumbai'),
        ('Annual Fee', '₹2-3 Lakhs')
    ]

    row += 1
    for label, value in school_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if isinstance(value, str) and not value.startswith('₹'):
            apply_input_style(ws[f'B{row}'], colors)
        row += 1

    # Challenge Selection
    row += 1
    ws[f'A{row}'] = 'CHALLENGE SELECTION (USER SELECTED 3)'
    apply_subheader_style(ws[f'A{row}'], colors)
    ws.merge_cells(f'A{row}:D{row}')

    challenges = [
        ('C1', 'Enrollment Decline', 50, 'Primary (50% weight)'),
        ('C4', 'Teacher Attrition', 30, 'Secondary (30% weight)'),
        ('C10', 'Parent Communication', 20, 'Tertiary (20% weight)')
    ]

    row += 1
    for cid, name, weight, description in challenges:
        ws[f'A{row}'] = cid
        ws[f'B{row}'] = name
        ws[f'C{row}'] = weight
        ws[f'D{row}'] = description
        ws[f'C{row}'].fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type='solid')
        row += 1

    # Screening Answers
    row += 1
    ws[f'A{row}'] = 'SCREENING QUESTION ANSWERS'
    apply_subheader_style(ws[f'A{row}'], colors)
    ws.merge_cells(f'A{row}:D{row}')

    answers = [
        ('C1-Q1.1', 'Enrollment Decline - Conversion Rate', 'Below 10%', 10),
        ('C1-Q1.2', 'Enrollment Decline - Marketing', 'No formal budget', 8),
        ('C1-Q1.3', 'Enrollment Decline - Drop-off', 'After inquiry', 9),
        ('C4-Q4.1', 'Teacher Attrition - Turnover Rate', '25%', 6),
        ('C4-Q4.2', 'Teacher Attrition - Teaching Load', '24-28 periods', 6),
        ('C4-Q4.3', 'Teacher Attrition - Exit Reason', 'Burnout', 9),
        ('C10-Q10.1', 'Parent Comm - Channel', 'WhatsApp/calls', 9),
        ('C10-Q10.2', 'Parent Comm - Response Time', 'Over 48 hours', 9),
    ]

    row += 1
    for qid, question, answer, weight in answers:
        ws[f'A{row}'] = qid
        ws[f'B{row}'] = question
        ws[f'C{row}'] = answer
        ws[f'D{row}'] = weight
        ws[f'D{row}'].fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type='solid')
        ws[f'D{row}'].font = Font(bold=True)
        row += 1

    # Operational Metrics
    row += 1
    ws[f'A{row}'] = 'OPERATIONAL METRICS INPUT'
    apply_subheader_style(ws[f'A{row}'], colors)
    ws.merge_cells(f'A{row}:D{row}')

    metrics = [
        ('Student-Teacher Ratio (STR)', 34, 'Students / Full-time Teachers'),
        ('Parent Response SLA (hours)', 52, 'Avg. hours to resolve query'),
        ('Teacher Training (hours/year)', 8, 'Annual Professional Development'),
        ('Weekly Planning Time (hours)', 2, 'Hours per teacher per week')
    ]

    row += 1
    for metric, value, desc in metrics:
        ws[f'A{row}'] = metric
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'C{row}'] = desc
        apply_input_style(ws[f'B{row}'], colors)
        row += 1

    # Store metric row numbers for formulas
    str_row = row - 4
    sla_row = row - 3
    train_row = row - 2
    plan_row = row - 1

    # Calculations
    row += 1
    ws[f'A{row}'] = 'CALCULATION STEPS'
    apply_subheader_style(ws[f'A{row}'], colors)
    ws.merge_cells(f'A{row}:D{row}')

    # Step 1: Subjective Base Score
    row += 1
    ws[f'A{row}'] = 'STEP 1: SUBJECTIVE BASE SCORE (S_sub)'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'A{row}'] = 'Formula:'
    ws[f'B{row}'] = '=100 - ((SUM(Question Weights) / (N × 10)) × 100)'
    ws[f'C{row}'] = 'Sum of all answer weights'

    row += 1
    ws[f'A{row}'] = 'Sum of weights:'
    ws[f'B{row}'] = 10+8+9+6+6+9+9+9  # Hardcoded sum
    ws[f'B{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'Number of questions:'
    ws[f'B{row}'] = 8
    ws[f'B{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'S_sub RESULT:'
    ws[f'B{row}'] = f'=100 - (({10+8+9+6+6+9+9+9})/(8*10))*100'
    apply_calc_style(ws[f'B{row}'], colors)
    ws[f'B{row}'].value = 100 - ((10+8+9+6+6+9+9+9)/(8*10))*100

    row += 1
    ws[f'A{row}'] = 'Interpretation:'
    ws[f'B{row}'] = 'School rates itself very poorly (low confidence)'
    ws.merge_cells(f'B{row}:D{row}')

    # Step 2: Objective Multipliers
    row += 2
    ws[f'A{row}'] = 'STEP 2: OBJECTIVE SCALING MULTIPLIERS'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'A{row}'] = 'STR = 34:1'
    ws[f'B{row}'] = '=IF(STR<=20, 1.05, IF(STR<=28, 1.00, IF(STR<=35, 0.88, 0.75)))'
    ws[f'C{row}'] = 'Overcrowded classrooms'
    ws[f'B{row}'].value = 0.88
    apply_calc_style(ws[f'B{row}'], colors)

    row += 1
    ws[f'A{row}'] = 'SLA = 52 hours'
    ws[f'B{row}'] = '=IF(SLA<=12, 1.05, IF(SLA<=24, 1.00, IF(SLA<=48, 0.85, 0.70)))'
    ws[f'C{row}'] = 'Severe communication breakdown'
    ws[f'B{row}'].value = 0.70
    apply_calc_style(ws[f'B{row}'], colors)

    row += 1
    ws[f'A{row}'] = 'Training = 8 hours/year'
    ws[f'B{row}'] = '=IF(Training>=25, 1.05, IF(Training>=15, 1.00, 0.85))'
    ws[f'C{row}'] = 'Stagnant pedagogy'
    ws[f'B{row}'].value = 0.85
    apply_calc_style(ws[f'B{row}'], colors)

    row += 1
    ws[f'A{row}'] = 'Planning = 2 hours/week'
    ws[f'B{row}'] = '=IF(Planning>=5, 1.05, IF(Planning>=3, 1.00, 0.88))'
    ws[f'C{row}'] = 'Ad-hoc classroom delivery'
    ws[f'B{row}'].value = 0.88
    apply_calc_style(ws[f'B{row}'], colors)

    row += 1
    ws[f'A{row}'] = 'M_obj RESULT:'
    ws[f'B{row}'] = '=0.88 × 0.70 × 0.85 × 0.88'
    ws[f'C{row}'] = 'Composite objective factor'
    ws[f'B{row}'].value = 0.88 * 0.70 * 0.85 * 0.88
    apply_calc_style(ws[f'B{row}'], colors)

    # Step 3: Scaled Score
    row += 2
    ws[f'A{row}'] = 'STEP 3: SCALED SCORE'
    ws[f'B{row}'] = '=S_sub × M_obj'
    ws[f'C{row}'] = 'Perception × Reality'

    row += 1
    s_sub_val = 100 - ((10+8+9+6+6+9+9+9)/(8*10))*100
    m_obj_val = 0.88 * 0.70 * 0.85 * 0.88
    scaled_val = s_sub_val * m_obj_val

    ws[f'A{row}'] = 'Scaled Score:'
    ws[f'B{row}'] = f'={s_sub_val:.2f} × {m_obj_val:.4f}'
    ws[f'C{row}'] = 'Severity × Operational Reality'
    ws[f'B{row}'].value = scaled_val
    apply_calc_style(ws[f'B{row}'], colors)

    # Step 4: Delusion Penalty
    row += 2
    ws[f'A{row}'] = 'STEP 4: DELUSION PENALTY'
    ws[f'B{row}'] = '=IF(AND(S_sub>=80, M_obj<=0.85), 15, IF(AND(S_sub>=70, M_obj<=0.78), 10, 0))'
    ws.merge_cells(f'B{row}:D{row}')

    row += 1
    ws[f'A{row}'] = 'Condition check:'
    ws[f'B{row}'] = f'S_sub ({s_sub_val:.2f}) >= 80? NO'
    ws[f'C{row}'] = 'No delusional comfort detected'

    row += 1
    ws[f'A{row}'] = 'Delusion Penalty:'
    ws[f'B{row}'] = 0
    apply_calc_style(ws[f'B{row}'], colors)

    # Final Health Index
    row += 2
    ws[f'A{row}'] = 'FINAL HEALTH INDEX (H)'
    apply_subheader_style(ws[f'A{row}'], colors)
    ws.merge_cells(f'A{row}:D{row}')
    ws.row_dimensions[row].height = 20

    row += 1
    ws[f'A{row}'] = 'Formula:'
    ws[f'B{row}'] = '=MAX(0, MIN(100, Scaled_Score - P_mismatch))'
    ws.merge_cells(f'B{row}:D{row}')

    row += 1
    ws[f'A{row}'] = 'Health Index:'
    health_index = max(0, min(100, scaled_val - 0))
    ws[f'B{row}'] = health_index
    apply_result_style(ws[f'B{row}'], colors)
    ws[f'B{row}'].font = Font(bold=True, size=14, color='FFDC2626')

    row += 1
    ws[f'A{row}'] = 'Risk Level:'
    if health_index >= 80:
        risk = 'LOW RISK'
        ws[f'B{row}'].fill = PatternFill(start_color='FF10B981', end_color='FF10B981', fill_type='solid')
    elif health_index >= 60:
        risk = 'MODERATE RISK'
        ws[f'B{row}'].fill = PatternFill(start_color='FFF59E0B', end_color='FFF59E0B', fill_type='solid')
    else:
        risk = 'HIGH RISK - RED ALERT'
        ws[f'B{row}'].fill = PatternFill(start_color='FFDC2626', end_color='FFDC2626', fill_type='solid')

    ws[f'B{row}'] = risk
    ws[f'B{row}'].font = Font(bold=True, color='FFFFFFFF', size=12)

    row += 1
    ws[f'A{row}'] = 'Risk Quadrant:'
    if s_sub_val >= 80 and m_obj_val >= 0.95:
        quadrant = 'ELITE EQUILIBRIUM'
    elif s_sub_val >= 80 and m_obj_val < 0.85:
        quadrant = 'DELUSIONAL COMFORT'
    elif s_sub_val < 60 and m_obj_val >= 0.95:
        quadrant = 'HIDDEN EXCELLENCE'
    else:
        quadrant = 'CRITICAL COLLAPSE'

    ws[f'B{row}'] = quadrant
    apply_result_style(ws[f'B{row}'], colors)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20

def main():
    print("Creating comprehensive First Opinion Engine Excel workbook...")

    wb = Workbook()
    wb.remove(wb.active)

    colors = create_color_scheme()

    # Create all sheets
    create_workflow_sheet(wb, colors)
    print("[OK] Sheet 1: User Workflow Overview")

    create_challenge_catalog(wb, colors)
    print("[OK] Sheet 2: 15 Challenges Catalog")

    create_questionnaire_master(wb, colors)
    print("[OK] Sheet 3: Master Questionnaire")

    create_example_scenario(wb, colors)
    print("[OK] Sheet 4: Complete Calculation Example")

    # Save workbook
    filename = 'public/DISHA_First_Opinion_Complete_Workflow.xlsx'
    wb.save(filename)

    print("\n" + "="*70)
    print("SUCCESS! COMPREHENSIVE FIRST OPINION ENGINE WORKBOOK CREATED")
    print("="*70)
    print(f"\nFile saved: {filename}")
    print("\nWorkbook contains:")
    print("  Sheet 1: User Workflow Overview (7 steps)")
    print("  Sheet 2: 15 Challenges Catalog (all challenges with metrics)")
    print("  Sheet 3: Master Questionnaire (all questions for 3 example challenges)")
    print("  Sheet 4: Live Calculation Example (hardcoded formulas & data mapping)")
    print("\nFeatures:")
    print("  • Color-coded sections for easy understanding")
    print("  • All formulas shown with hardcoded values")
    print("  • Complete data mapping from inputs to calculations")
    print("  • Example scenario with 3 selected challenges")
    print("  • Step-by-step calculation flow")
    print("  • Risk quadrant classification")
    print("\nReady to open in Microsoft Excel!")

if __name__ == '__main__':
    main()
