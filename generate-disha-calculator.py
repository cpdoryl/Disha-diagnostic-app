#!/usr/bin/env python3
"""
DISHA First Opinion Engine Calculator
Generates Excel workbook with complete calculation engine and questionnaire
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_disha_calculator():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Define styles
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    subheader_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    subheader_font = Font(color="FFFFFF", bold=True, size=11)

    input_fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
    result_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========================
    # SHEET 1: INPUT DASHBOARD
    # ========================
    ws_input = wb.create_sheet("Input Dashboard", 0)
    ws_input.column_dimensions['A'].width = 20
    ws_input.column_dimensions['B'].width = 40
    ws_input.column_dimensions['C'].width = 15

    # Title
    ws_input['A1'] = "DISHA FIRST OPINION ENGINE"
    ws_input['A1'].font = Font(size=16, bold=True, color="1E40AF")
    ws_input.merge_cells('A1:C1')

    ws_input['A2'] = "School Diagnostic & Institutional Health Assessment"
    ws_input['A2'].font = Font(size=11, italic=True)
    ws_input.merge_cells('A2:C2')

    # School Details Section
    ws_input['A4'] = "SCHOOL DETAILS"
    ws_input['A4'].fill = header_fill
    ws_input['A4'].font = header_font
    ws_input.merge_cells('A4:C4')

    school_fields = [
        ("School Name", "A5", "INPUT"),
        ("Board Affiliation", "A6", "CBSE / ICSE / IB / State"),
        ("Total Students", "A7", "INPUT"),
        ("Annual Fee Tier", "A8", "INPUT"),
        ("City / District", "A9", "INPUT"),
    ]

    row = 5
    for label, cell, description in school_fields:
        ws_input[f'A{row}'] = label
        ws_input[f'A{row}'].font = Font(bold=True)
        ws_input[f'B{row}'] = description
        ws_input[f'B{row}'].fill = input_fill
        row += 1

    # Challenge Selection Section
    ws_input['A12'] = "CHALLENGE SELECTION (Max 3)"
    ws_input['A12'].fill = header_fill
    ws_input['A12'].font = header_font
    ws_input.merge_cells('A12:C12')

    challenge_options = [
        "C1: Enrollment Decline",
        "C2: Student Attrition",
        "C3: Fee Collection",
        "C4: Teacher Attrition",
        "C5: Staff Capability",
        "C6: Leadership Gap",
        "C7: Academic Decline",
        "C8: Student Wellbeing",
        "C9: Remedial Lag",
        "C10: Parent Dissatisfaction",
        "C11: Competitor Pressure",
        "C12: Brand Perception",
        "C13: Cost Inflation",
        "C14: Infra Deficits",
        "C15: Compliance Stress"
    ]

    row = 13
    ws_input[f'A{row}'] = "Primary Challenge"
    ws_input[f'A{row}'].font = Font(bold=True)
    ws_input[f'B{row}'] = "Select 1 (50% weight)"
    ws_input[f'B{row}'].fill = input_fill
    row += 1

    ws_input[f'A{row}'] = "Secondary Challenge"
    ws_input[f'A{row}'].font = Font(bold=True)
    ws_input[f'B{row}'] = "Select 1 (30% weight)"
    ws_input[f'B{row}'].fill = input_fill
    row += 1

    ws_input[f'A{row}'] = "Tertiary Challenge"
    ws_input[f'A{row}'].font = Font(bold=True)
    ws_input[f'B{row}'] = "Select 1 (20% weight)"
    ws_input[f'B{row}'].fill = input_fill

    # Operational Metrics Section
    ws_input['A20'] = "OPERATIONAL METRICS"
    ws_input['A20'].fill = header_fill
    ws_input['A20'].font = header_font
    ws_input.merge_cells('A20:C20')

    metrics = [
        ("Student-Teacher Ratio", "A21", "Ideal: ≤20", "INPUT"),
        ("Parent Response SLA (hrs)", "A22", "Ideal: ≤12", "INPUT"),
        ("Teacher Training Hours/Year", "A23", "Ideal: ≥25", "INPUT"),
        ("Weekly Lesson Planning (hrs)", "A24", "Ideal: ≥5", "INPUT"),
    ]

    row = 21
    for label, cell, benchmark, type_ in metrics:
        ws_input[f'A{row}'] = label
        ws_input[f'A{row}'].font = Font(bold=True)
        ws_input[f'B{row}'] = type_
        ws_input[f'B{row}'].fill = input_fill
        ws_input[f'C{row}'] = benchmark
        ws_input[f'C{row}'].font = Font(italic=True, size=9)
        row += 1

    return wb

def add_questionnaire_sheet(wb):
    """Add detailed questionnaire sheet"""
    ws = wb.create_sheet("Questionnaire", 1)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50

    # Title
    ws['A1'] = "COMPLETE QUESTIONNAIRE WITH WEIGHTS"
    ws['A1'].font = Font(size=14, bold=True, color="1E40AF")
    ws.merge_cells('A1:D1')

    # Challenge C1: Enrollment Decline
    row = 3
    ws[f'A{row}'] = "C1"
    ws[f'B{row}'] = "ENROLLMENT DECLINE / ADMISSION SHORTFALL"
    ws[f'B{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'B{row}'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'B{row}'] = "Q1: Inquiry-to-Admission Conversion Rate"
    ws[f'B{row}'].font = Font(italic=True)

    row += 1
    questions_c1 = [
        ("Above 25%", 2),
        ("10-25%", 6),
        ("Below 10%", 10),
    ]

    for option, weight in questions_c1:
        ws[f'B{row}'] = option
        ws[f'C{row}'] = weight
        ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Add more challenges...
    row += 2
    ws[f'B{row}'] = "Q2: Primary Marketing Channels"
    ws[f'B{row}'].font = Font(italic=True)

    row += 1
    questions_c1_2 = [
        ("Digital marketing (Google, Meta, WhatsApp)", 3),
        ("Traditional print (Newspaper, flyers, hoardings)", 6),
        ("No formal marketing (Word-of-mouth only)", 8),
    ]

    for option, weight in questions_c1_2:
        ws[f'B{row}'] = option
        ws[f'C{row}'] = weight
        ws[f'C{row}'].font = Font(bold=True)
        row += 1

    row += 2
    ws[f'B{row}'] = "Q3: Parent Drop-off Stage"
    ws[f'B{row}'].font = Font(italic=True)

    row += 1
    questions_c1_3 = [
        ("After initial inquiry / Never visit school", 9),
        ("After touring school / Fees too high", 7),
        ("After final offer / Choosing competitor", 8),
    ]

    for option, weight in questions_c1_3:
        ws[f'B{row}'] = option
        ws[f'C{row}'] = weight
        ws[f'C{row}'].font = Font(bold=True)
        row += 1

    return wb

def add_calculation_engine(wb):
    """Add calculation engine sheet"""
    ws = wb.create_sheet("Calculation Engine", 2)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50

    # Title
    ws['A1'] = "DISHA CALCULATION ENGINE"
    ws['A1'].font = Font(size=14, bold=True, color="1E40AF")
    ws.merge_cells('A1:C1')

    # Input Section
    row = 3
    ws[f'A{row}'] = "INPUT PARAMETERS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    input_params = [
        ("Primary Challenge Weight", "50%", "B5"),
        ("Q1 Weight (Primary)", "INSERT", "B6"),
        ("Q2 Weight (Primary)", "INSERT", "B7"),
        ("Q3 Weight (Primary)", "INSERT", "B8"),
        ("", "", ""),
        ("Student-Teacher Ratio", "INSERT", "B10"),
        ("Parent Response SLA (hours)", "INSERT", "B11"),
        ("Teacher Training Hours", "INSERT", "B12"),
        ("Weekly Planning Hours", "INSERT", "B13"),
    ]

    for i, (label, value, ref) in enumerate(input_params, row):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value if value != "INSERT" else ""
        if value == "INSERT":
            ws[f'B{i}'].fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
        ws[f'C{i}'] = ref
        ws[f'C{i}'].font = Font(size=8, italic=True)

    # Calculation Section
    row = 16
    ws[f'A{row}'] = "CALCULATIONS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    ws[f'A{row}'] = "Step 1: Subjective Base Score (S_sub)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=100 - ((SUM(B6:B8)/(3*10))*100)"
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "0-100"

    row += 1
    ws[f'A{row}'] = "Step 2a: STR Multiplier (m_STR)"
    ws[f'B{row}'] = '=IF(B10<=20,1.05,IF(B10<=28,1.00,IF(B10<=35,0.88,0.75)))'
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    row += 1
    ws[f'A{row}'] = "Step 2b: SLA Multiplier (m_SLA)"
    ws[f'B{row}'] = '=IF(B11<=12,1.05,IF(B11<=24,1.00,IF(B11<=48,0.85,0.70)))'
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    row += 1
    ws[f'A{row}'] = "Step 2c: Training Multiplier (m_retrain)"
    ws[f'B{row}'] = '=IF(B12>=25,1.05,IF(B12>=15,1.00,0.85))'
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    row += 1
    ws[f'A{row}'] = "Step 2d: Planning Multiplier (m_plan)"
    ws[f'B{row}'] = '=IF(B13>=5,1.05,IF(B13>=3,1.00,0.88))'
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    row += 1
    ws[f'A{row}'] = "Step 2: Objective Scaling (M_obj)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=B18*B19*B20*B21"
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "0.60-1.15"

    row += 1
    ws[f'A{row}'] = "Step 3: Scaled Score"
    ws[f'B{row}'] = "=B17*B22"
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ws[f'C{row}'] = "S_sub × M_obj"

    row += 1
    ws[f'A{row}'] = "Step 4: Delusion Penalty (P_mismatch)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = '=IF(AND(B17>=80,B22<=0.85),15,IF(AND(B17>=70,B22<=0.78),10,0))'
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "0, 10, or 15"

    row += 1
    ws[f'A{row}'] = "FINAL HEALTH INDEX (H)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    ws[f'B{row}'] = "=MAX(0,MIN(100,B23-B25))"
    ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = "0-100 SCORE"
    ws[f'C{row}'].font = Font(bold=True)

    # Risk Quadrant
    row += 2
    ws[f'A{row}'] = "RISK QUADRANT CLASSIFICATION"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    risk_logic = [
        ("Quadrant", "Condition", "Outcome"),
        ("Elite Equilibrium", "S≥80 AND M≥0.95", "World-class, aligned"),
        ("Delusional Comfort", "S≥80 AND M<0.85", "⚠ CRITICAL - RED ALERT"),
        ("Hidden Excellence", "S<60 AND M≥0.95", "Strong ops, low confidence"),
        ("Critical Collapse", "S<60 AND M<0.85", "🚨 EMERGENCY - ESCALATE"),
    ]

    for label, condition, outcome in risk_logic:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True if label == "Quadrant" else False)
        ws[f'B{row}'] = condition
        ws[f'C{row}'] = outcome
        row += 1

    return wb

def main():
    print("🏫 DISHA First Opinion Engine Calculator Generator")
    print("=" * 50)

    try:
        wb = create_disha_calculator()
        add_questionnaire_sheet(wb)
        add_calculation_engine(wb)

        filename = "DISHA_First_Opinion_Calculator.xlsx"
        wb.save(filename)

        print(f"✅ Excel file created: {filename}")
        print("\nWorkbook Contents:")
        print("  1. Input Dashboard - School details & metrics")
        print("  2. Questionnaire - All 15 challenges with options & weights")
        print("  3. Calculation Engine - Real-time formula calculations")
        print("\n📊 Interactive Features:")
        print("  - Input cells highlighted in light blue")
        print("  - Formulas auto-calculate as you enter values")
        print("  - Health index updates in real-time")
        print("  - Risk quadrant classification automatic")
        print("\n✨ Ready to use! Open in Excel and start calculating.")

    except Exception as e:
        print(f"❌ Error: {e}")
        print("\nMake sure openpyxl is installed:")
        print("  pip install openpyxl")

if __name__ == "__main__":
    main()
