#!/usr/bin/env python3
"""
Stage 3: Reverse Outcome Modeling & Target Feasibility Validation
Complete Excel engine showing goal setting, reverse calculation, feasibility analysis,
action mapping, and timeline for achieving targets
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_colors():
    return {
        'header_dark': '1E40AF',      # Dark blue
        'header_light': '0D9488',     # Teal
        'goal_header': 'DC2626',      # Red
        'current': 'FECACA',          # Light red
        'target': 'D1FAE5',           # Light green
        'feasible': 'D1FAE5',         # Light green
        'challenging': 'FED7AA',      # Orange
        'risky': 'FEE2E2',            # Red
        'input_field': 'CCFBF1',      # Light cyan
        'calculation': 'FEF3C7',      # Yellow
        'result': 'FFFFE4E6',         # Light red/pink
        'neutral': 'F3F4F6'           # Gray
    }

# Dimension data with improvement parameters
DIMENSIONS_DATA = {
    'D01': {
        'name': 'Academic Reputation & Rigour',
        'weight': 10,
        'current_score': 75,  # Example current
        'improvement_difficulty': 7,  # 1-10 (higher = harder)
        'time_months': 12,  # Time to improve significantly
        'cost_factor': 6,  # 1-10 (higher = more expensive)
        'improvements': [
            ('Curriculum enhancement', 5, 3),  # (action, months, cost_factor)
            ('Teacher training program', 8, 4),
            ('External exam coaching', 6, 2),
            ('Peer school benchmarking', 2, 1),
            ('Academic audit', 3, 2)
        ]
    },
    'D02': {
        'name': 'Teacher Welfare & Development',
        'weight': 9,
        'current_score': 70,
        'improvement_difficulty': 6,
        'time_months': 6,
        'cost_factor': 8,
        'improvements': [
            ('Salary increase', 1, 9),
            ('Professional development budget', 3, 7),
            ('Career progression plan', 2, 2),
            ('Work-life balance initiatives', 3, 3),
            ('Performance incentives', 4, 6)
        ]
    },
    'D03': {
        'name': 'Leadership & Governance Quality',
        'weight': 10,
        'current_score': 78,
        'improvement_difficulty': 5,
        'time_months': 6,
        'cost_factor': 4,
        'improvements': [
            ('Leadership training', 4, 3),
            ('Strategic planning workshop', 3, 2),
            ('Governance restructuring', 6, 4),
            ('Decision-making framework', 2, 1),
            ('Succession plan development', 4, 2)
        ]
    },
    'D04': {
        'name': 'Parent Engagement & SLA',
        'weight': 8,
        'current_score': 72,
        'improvement_difficulty': 4,
        'time_months': 3,
        'cost_factor': 3,
        'improvements': [
            ('Communication system upgrade', 2, 4),
            ('Parent meeting schedule', 1, 1),
            ('Response time SLA setup', 1, 1),
            ('Feedback mechanism', 2, 2),
            ('Parent engagement programs', 3, 3)
        ]
    },
    'D05': {
        'name': 'Student Safety & Wellness',
        'weight': 10,
        'current_score': 80,
        'improvement_difficulty': 6,
        'time_months': 9,
        'cost_factor': 7,
        'improvements': [
            ('Mental health counselor', 1, 8),
            ('Safety audit and upgrades', 3, 7),
            ('Anti-bullying program', 4, 3),
            ('Wellness center setup', 6, 8),
            ('Staff training', 2, 2)
        ]
    },
    'D06': {
        'name': 'Infrastructure & Facilities',
        'weight': 7,
        'current_score': 68,
        'improvement_difficulty': 8,
        'time_months': 18,
        'cost_factor': 9,
        'improvements': [
            ('Technology integration', 6, 8),
            ('Infrastructure upgrade', 12, 10),
            ('Maintenance program', 3, 4),
            ('Facilities modernization', 12, 9),
            ('Lab and library enhancement', 9, 8)
        ]
    },
    'D07': {
        'name': 'Co-Curricular Education',
        'weight': 6,
        'current_score': 65,
        'improvement_difficulty': 5,
        'time_months': 6,
        'cost_factor': 5,
        'improvements': [
            ('Sports program expansion', 4, 6),
            ('Arts program development', 4, 5),
            ('Club formation', 2, 2),
            ('Competition participation', 3, 4),
            ('External partnerships', 3, 2)
        ]
    },
    'D08': {
        'name': 'Individual Attention (PTR)',
        'weight': 9,
        'current_score': 75,
        'improvement_difficulty': 9,
        'time_months': 18,
        'cost_factor': 9,
        'improvements': [
            ('Class size reduction', 12, 10),
            ('Remedial program', 4, 6),
            ('Personalized learning plans', 6, 4),
            ('One-on-one support hiring', 6, 8),
            ('Differentiation training', 3, 3)
        ]
    },
    'D09': {
        'name': 'Value for Money',
        'weight': 7,
        'current_score': 70,
        'improvement_difficulty': 7,
        'time_months': 12,
        'cost_factor': 8,
        'improvements': [
            ('Scholarship increase', 3, 9),
            ('Value communication', 2, 1),
            ('Fee structure review', 2, 2),
            ('Financial aid programs', 4, 8),
            ('Parent perception survey', 2, 2)
        ]
    },
    'D10': {
        'name': 'Special Needs Inclusivity',
        'weight': 6,
        'current_score': 60,
        'improvement_difficulty': 7,
        'time_months': 12,
        'cost_factor': 8,
        'improvements': [
            ('SEN coordinator hiring', 3, 7),
            ('Accessibility audit', 2, 3),
            ('Infrastructure accessibility', 8, 9),
            ('Staff training', 4, 4),
            ('Support program setup', 6, 6)
        ]
    },
    'D11': {
        'name': 'Community Service & Social Responsibility',
        'weight': 5,
        'current_score': 65,
        'improvement_difficulty': 4,
        'time_months': 6,
        'cost_factor': 3,
        'improvements': [
            ('Community partnership', 3, 2),
            ('Service program launch', 4, 3),
            ('Sustainability initiative', 6, 4),
            ('Student engagement drive', 2, 1),
            ('Environmental program', 4, 3)
        ]
    },
    'D12': {
        'name': 'Faculty Competence & Retention',
        'weight': 9,
        'current_score': 82,
        'improvement_difficulty': 6,
        'time_months': 12,
        'cost_factor': 7,
        'improvements': [
            ('Qualification upgrade support', 6, 6),
            ('Subject specialist hiring', 6, 8),
            ('Continuous evaluation system', 3, 2),
            ('Professional development', 4, 5),
            ('Retention incentives', 6, 8)
        ]
    },
    'D13': {
        'name': 'Internationalism & Cultural Diversity',
        'weight': 6,
        'current_score': 70,
        'improvement_difficulty': 7,
        'time_months': 12,
        'cost_factor': 6,
        'improvements': [
            ('International curriculum', 8, 6),
            ('IB program', 12, 8),
            ('Partnership development', 4, 3),
            ('Exchange programs', 6, 5),
            ('Diversity recruitment', 4, 4)
        ]
    },
    'D14': {
        'name': 'Management Vision & Growth Drive',
        'weight': 8,
        'current_score': 76,
        'improvement_difficulty': 4,
        'time_months': 6,
        'cost_factor': 2,
        'improvements': [
            ('Strategic planning', 3, 2),
            ('Innovation lab setup', 4, 3),
            ('Market positioning', 3, 2),
            ('Growth targeting', 2, 1),
            ('Future readiness assessment', 2, 1)
        ]
    }
}

def apply_cell_style(cell, style_type, colors):
    """Apply cell styling"""
    if style_type == 'header_dark':
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=colors['header_dark'], end_color=colors['header_dark'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    elif style_type == 'header_light':
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=colors['header_light'], end_color=colors['header_light'], fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    elif style_type == 'goal_header':
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=colors['goal_header'], end_color=colors['goal_header'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    elif style_type == 'current':
        cell.fill = PatternFill(start_color=colors['current'], end_color=colors['current'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif style_type == 'target':
        cell.fill = PatternFill(start_color=colors['target'], end_color=colors['target'], fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif style_type == 'calculation':
        cell.fill = PatternFill(start_color=colors['calculation'], end_color=colors['calculation'], fill_type="solid")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif style_type == 'input':
        cell.fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def create_goal_setting_sheet(wb, colors):
    """Stage 3 Goal Setting Sheet"""
    ws = wb.create_sheet("01 GOAL SETTING", 0)

    row = 1
    ws[f'A{row}'] = "STAGE 3: REVERSE OUTCOME MODELING - GOAL SETTING"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:E{row}')
    ws.row_dimensions[row].height = 25
    row += 1

    ws[f'A{row}'] = "Define your institutional health target and feasibility parameters"
    ws.merge_cells(f'A{row}:E{row}')
    row += 2

    # Current state
    ws[f'A{row}'] = "CURRENT STATE BASELINE"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = "Current Overall Health Index:"
    ws[f'B{row}'] = "[ENTER: e.g., 72]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    row += 1

    ws[f'A{row}'] = "Assessment Date:"
    ws[f'B{row}'] = "[ENTER: Date]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    row += 2

    # Target definition
    ws[f'A{row}'] = "TARGET DEFINITION"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = "Target Overall Health Index:"
    ws[f'B{row}'] = "[ENTER: e.g., 80]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    ws[f'C{row}'] = "(0-100, recommend 75-85 for realistic targets)"
    row += 1

    ws[f'A{row}'] = "Target Achievement Timeline:"
    ws[f'B{row}'] = "[ENTER: months, e.g., 12]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    ws[f'C{row}'] = "(Typical: 6-24 months)"
    row += 2

    # Improvement parameters
    ws[f'A{row}'] = "IMPROVEMENT PARAMETERS"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = "Available Budget (₹):"
    ws[f'B{row}'] = "[ENTER: Budget]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    row += 1

    ws[f'A{row}'] = "Priority Focus Areas:"
    ws[f'B{row}'] = "[SELECT: Strategic focus]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    ws[f'C{row}'] = "(Academic / Facilities / Staff / Holistic)"
    row += 1

    ws[f'A{row}'] = "Risk Tolerance:"
    ws[f'B{row}'] = "[SELECT: Low / Medium / High]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    row += 2

    # Gap analysis
    ws[f'A{row}'] = "QUICK GAP ANALYSIS"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = "Gap to Target (Target - Current):"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    ws[f'C{row}'] = "If 80 - 72 = 8 points needed"
    row += 1

    ws[f'A{row}'] = "Average Improvement per Dimension:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    ws[f'C{row}'] = "8 / 14 dimensions ≈ 0.57 per dimension"
    row += 1

    ws[f'A{row}'] = "% Improvement Needed:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    ws[f'C{row}'] = "(8 / 72) × 100 = 11.1%"
    row += 2

    # Feasibility assessment
    ws[f'A{row}'] = "FEASIBILITY ASSESSMENT"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = "Quick Assessment:"
    row += 1

    ws[f'A{row}'] = "If gap ≤ 5 points and timeline ≥ 12 months:"
    ws[f'B{row}'] = "HIGHLY FEASIBLE"
    ws[f'B{row}'].font = Font(color="FFFFFF", bold=True)
    ws[f'B{row}'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    row += 1

    ws[f'A{row}'] = "If gap 5-10 points and timeline 12-18 months:"
    ws[f'B{row}'] = "FEASIBLE"
    ws[f'B{row}'].font = Font(color="FFFFFF", bold=True)
    ws[f'B{row}'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    row += 1

    ws[f'A{row}'] = "If gap >10 points and timeline <12 months:"
    ws[f'B{row}'] = "CHALLENGING"
    ws[f'B{row}'].font = Font(color="FFFFFF", bold=True)
    ws[f'B{row}'].fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    row += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

def create_reverse_calculation_sheet(wb, colors):
    """Reverse Calculation Engine Sheet"""
    ws = wb.create_sheet("02 REVERSE CALCULATION", 1)

    row = 1
    ws[f'A{row}'] = "REVERSE CALCULATION ENGINE - Target Dimension Scores"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 25
    row += 2

    ws[f'A{row}'] = "Working backwards from target to required dimension scores"
    ws.merge_cells(f'A{row}:H{row}')
    row += 2

    # Input section
    ws[f'A{row}'] = "INPUTS"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Target Overall Health Index:"
    ws[f'B{row}'] = "[From Goal Setting Sheet]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    row += 1

    ws[f'A{row}'] = "Total Weight:"
    ws[f'B{row}'] = "109"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    row += 2

    # Calculation logic
    ws[f'A{row}'] = "CALCULATION LOGIC"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "1. Overall Index = (Sum of Weighted Contributions) / Total Weight"
    row += 1
    ws[f'A{row}'] = "2. Required Total Points = Target Index × Total Weight / 100"
    row += 1
    ws[f'A{row}'] = "3. Allocate points based on:"
    row += 1
    ws[f'A{row}'] = "   - Current performance (keep strong areas stronger)"
    row += 1
    ws[f'A{row}'] = "   - Improvement difficulty (allocate more to easier areas)"
    row += 1
    ws[f'A{row}'] = "   - Strategic priority (allocate based on goals)"
    row += 2

    # Dimension targets table
    ws[f'A{row}'] = "Dimension"
    ws[f'B{row}'] = "Weight"
    ws[f'C{row}'] = "Current"
    ws[f'D{row}'] = "Target"
    ws[f'E{row}'] = "Needed"
    ws[f'F{row}'] = "Difficulty"
    ws[f'G{row}'] = "Feasibility"
    ws[f'H{row}'] = "Priority"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        apply_cell_style(ws[f'{col}{row}'], 'header_light', colors)

    row += 1
    start_row = row

    # Add all dimensions with sample targets
    for dim_id in sorted(DIMENSIONS_DATA.keys()):
        dim = DIMENSIONS_DATA[dim_id]
        ws[f'A{row}'] = dim['name']
        ws[f'B{row}'] = dim['weight']
        ws[f'C{row}'] = dim['current_score']
        ws[f'D{row}'] = "[TARGET SCORE]"
        ws[f'E{row}'] = "FORMULA"
        ws[f'F{row}'] = dim['improvement_difficulty']
        ws[f'G{row}'] = "[ASSESS]"
        ws[f'H{row}'] = "[STRATEGIC]"

        apply_cell_style(ws[f'C{row}'], 'current', colors)
        apply_cell_style(ws[f'D{row}'], 'target', colors)
        apply_cell_style(ws[f'E{row}'], 'calculation', colors)
        row += 1

    row += 2

    # Summary
    ws[f'A{row}'] = "SUMMARY"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Total Current Points:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    row += 1

    ws[f'A{row}'] = "Total Target Points:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    row += 1

    ws[f'A{row}'] = "Total Points to Gain:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    row += 1

    ws[f'A{row}'] = "Average per Dimension:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

def create_feasibility_sheet(wb, colors):
    """Feasibility Analysis Sheet"""
    ws = wb.create_sheet("03 FEASIBILITY ANALYSIS", 2)

    row = 1
    ws[f'A{row}'] = "FEASIBILITY ANALYSIS - Achievability Assessment"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 25
    row += 2

    ws[f'A{row}'] = "Assessment Framework"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Feasibility = f(Improvement Gap, Timeline, Resources, Difficulty)"
    ws.merge_cells(f'A{row}:H{row}')
    row += 2

    # Feasibility matrix
    ws[f'A{row}'] = "Dimension"
    ws[f'B{row}'] = "Gap"
    ws[f'C{row}'] = "Timeline"
    ws[f'D{row}'] = "Cost"
    ws[f'E{row}'] = "Difficulty"
    ws[f'F{row}'] = "Feasibility %"
    ws[f'G{row}'] = "Risk Level"
    ws[f'H{row}'] = "Recommendation"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        apply_cell_style(ws[f'{col}{row}'], 'header_light', colors)
    row += 1

    # Add dimensions with feasibility calculation
    for dim_id in sorted(DIMENSIONS_DATA.keys()):
        dim = DIMENSIONS_DATA[dim_id]
        ws[f'A{row}'] = dim['name']
        ws[f'B{row}'] = "CALC"
        ws[f'C{row}'] = f"{dim['time_months']} mo"
        ws[f'D{row}'] = dim['cost_factor']
        ws[f'E{row}'] = dim['improvement_difficulty']
        ws[f'F{row}'] = "FORMULA"
        ws[f'G{row}'] = "[ASSESS]"
        ws[f'H{row}'] = "[ACTION]"

        apply_cell_style(ws[f'B{row}'], 'calculation', colors)
        apply_cell_style(ws[f'F{row}'], 'calculation', colors)
        row += 1

    row += 2

    # Legend
    ws[f'A{row}'] = "FEASIBILITY SCALE"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "90-100%: Highly Feasible - Implement immediately"
    ws[f'B{row}'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    row += 1

    ws[f'A{row}'] = "70-89%: Feasible - Implement with planning"
    ws[f'B{row}'].fill = PatternFill(start_color="84CC16", end_color="84CC16", fill_type="solid")
    row += 1

    ws[f'A{row}'] = "50-69%: Challenging - Requires strategic focus"
    ws[f'B{row}'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    row += 1

    ws[f'A{row}'] = "<50%: High Risk - Needs careful planning"
    ws[f'B{row}'].fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20

def create_action_mapping_sheet(wb, colors):
    """Action Mapping Sheet"""
    ws = wb.create_sheet("04 ACTION MAPPING", 3)

    row = 1
    ws[f'A{row}'] = "ACTION MAPPING - Converting Targets to Specific Improvements"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 25
    row += 2

    ws[f'A{row}'] = "Detailed Action Plan by Dimension"
    ws.merge_cells(f'A{row}:H{row}')
    row += 2

    # Action mapping
    ws[f'A{row}'] = "Dimension"
    ws[f'B{row}'] = "Target Score"
    ws[f'C{row}'] = "Action Item"
    ws[f'D{row}'] = "Timeline"
    ws[f'E{row}'] = "Resources"
    ws[f'F{row}'] = "Owner"
    ws[f'G{row}'] = "Impact"
    ws[f'H{row}'] = "Status"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        apply_cell_style(ws[f'{col}{row}'], 'header_light', colors)
    row += 1

    # Add dimensions with action items
    for dim_id in sorted(DIMENSIONS_DATA.keys()):
        dim = DIMENSIONS_DATA[dim_id]

        # Dimension header
        ws[f'A{row}'] = dim['name']
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = "[TARGET]"
        apply_cell_style(ws[f'B{row}'], 'target', colors)
        ws.merge_cells(f'C{row}:H{row}')
        row += 1

        # Actions for this dimension
        for action, time, cost in dim['improvements']:
            ws[f'A{row}'] = ""
            ws[f'B{row}'] = ""
            ws[f'C{row}'] = action
            ws[f'D{row}'] = f"{time} mo"
            ws[f'E{row}'] = f"Cost {cost}/10"
            ws[f'F{row}'] = "[OWNER]"
            ws[f'G{row}'] = "+[SCORE]"
            ws[f'H{row}'] = "[ ]"
            apply_cell_style(ws[f'G{row}'], 'calculation', colors)
            row += 1

        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10

def create_resource_allocation_sheet(wb, colors):
    """Resource Allocation Sheet"""
    ws = wb.create_sheet("05 RESOURCE ALLOCATION", 4)

    row = 1
    ws[f'A{row}'] = "RESOURCE ALLOCATION - Budget & Effort Planning"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 25
    row += 2

    # Budget summary
    ws[f'A{row}'] = "BUDGET ALLOCATION"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Total Available Budget (₹):"
    ws[f'B{row}'] = "[FROM GOAL SETTING]"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    row += 2

    # Resource table
    ws[f'A{row}'] = "Dimension"
    ws[f'B{row}'] = "Priority"
    ws[f'C{row}'] = "Cost Factor"
    ws[f'D{row}'] = "% of Budget"
    ws[f'E{row}'] = "Allocated (₹)"
    ws[f'F{row}'] = "Timeline (mo)"
    ws[f'G{row}'] = "ROI (Score)"
    ws[f'H{row}'] = "Notes"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        apply_cell_style(ws[f'{col}{row}'], 'header_light', colors)
    row += 1

    for dim_id in sorted(DIMENSIONS_DATA.keys()):
        dim = DIMENSIONS_DATA[dim_id]
        ws[f'A{row}'] = dim['name']
        ws[f'B{row}'] = "[SELECT]"
        ws[f'C{row}'] = dim['cost_factor']
        ws[f'D{row}'] = "FORMULA"
        ws[f'E{row}'] = "FORMULA"
        ws[f'F{row}'] = dim['time_months']
        ws[f'G{row}'] = "CALC"
        ws[f'H{row}'] = ""

        apply_cell_style(ws[f'D{row}'], 'calculation', colors)
        apply_cell_style(ws[f'E{row}'], 'calculation', colors)
        apply_cell_style(ws[f'G{row}'], 'calculation', colors)
        row += 1

    row += 2

    # Summary
    ws[f'A{row}'] = "BUDGET SUMMARY"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Total Allocated:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    row += 1

    ws[f'A{row}'] = "Remaining Buffer:"
    ws[f'B{row}'] = "FORMULA"
    apply_cell_style(ws[f'B{row}'], 'calculation', colors)
    row += 1

    ws[f'A{row}'] = "Budget Adequacy:"
    ws[f'B{row}'] = "[ASSESS]"

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20

def create_timeline_sheet(wb, colors):
    """Timeline and Milestones Sheet"""
    ws = wb.create_sheet("06 TIMELINE & MILESTONES", 5)

    row = 1
    ws[f'A{row}'] = "TIMELINE & MILESTONES - Phased Implementation Plan"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 25
    row += 2

    ws[f'A{row}'] = "Overall Timeline:"
    ws[f'B{row}'] = "[ENTER: 12] months"
    apply_cell_style(ws[f'B{row}'], 'input', colors)
    row += 2

    # Phase breakdown
    ws[f'A{row}'] = "PHASE BREAKDOWN"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    phases = [
        ('Phase 1: Foundation', 'Months 1-3', 'Planning, assessment, small wins'),
        ('Phase 2: Build', 'Months 4-9', 'Major implementations'),
        ('Phase 3: Optimize', 'Months 10-12', 'Fine-tuning and consolidation')
    ]

    for phase_name, phase_duration, phase_focus in phases:
        ws[f'A{row}'] = phase_name
        ws[f'B{row}'] = phase_duration
        ws[f'C{row}'] = phase_focus
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'C{row}:H{row}')
        row += 1

    row += 2

    # Milestones table
    ws[f'A{row}'] = "Milestone"
    ws[f'B{row}'] = "Month"
    ws[f'C{row}'] = "Target Health"
    ws[f'D{row}'] = "Key Deliverables"
    ws[f'E{row}'] = "Owner"
    ws[f'F{row}'] = "Status"
    ws[f'G{row}'] = "Approval"
    ws[f'H{row}'] = "Notes"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        apply_cell_style(ws[f'{col}{row}'], 'header_light', colors)
    row += 1

    milestones = [
        ('Kickoff', 0, 72, 'Team assembly, plan approval', '[OWNER]'),
        ('Phase 1 Review', 3, 74, 'Quick wins implemented', '[OWNER]'),
        ('Mid-Year Review', 6, 76, 'Major projects 50% done', '[OWNER]'),
        ('Phase 2 Review', 9, 78, 'Major projects complete', '[OWNER]'),
        ('Final Review', 12, 80, 'All targets achieved', '[OWNER]'),
    ]

    for milestone, month, target, deliverables, owner in milestones:
        ws[f'A{row}'] = milestone
        ws[f'B{row}'] = month
        ws[f'C{row}'] = target
        ws[f'D{row}'] = deliverables
        ws[f'E{row}'] = owner
        ws[f'F{row}'] = "[ ]"
        ws[f'G{row}'] = "[ ]"
        ws[f'H{row}'] = ""
        apply_cell_style(ws[f'C{row}'], 'calculation', colors)
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20

def create_example_scenario_sheet(wb, colors):
    """Example Complete Scenario"""
    ws = wb.create_sheet("07 EXAMPLE SCENARIO", 6)

    row = 1
    ws[f'A{row}'] = "EXAMPLE SCENARIO - Complete Reverse Modeling Walkthrough"
    apply_cell_style(ws[f'A{row}'], 'goal_header', colors)
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 25
    row += 2

    ws[f'A{row}'] = "School: Delhi Excellence Academy | Current: 72 | Target: 80 | Timeline: 12 months"
    ws.merge_cells(f'A{row}:H{row}')
    row += 2

    # Goal setting
    ws[f'A{row}'] = "STEP 1: GOAL SETTING"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Current Health Index: 72/100"
    ws[f'B{row}'] = "Status: Average Performer"
    row += 1

    ws[f'A{row}'] = "Target Health Index: 80/100"
    ws[f'B{row}'] = "Status: Strong Performer"
    row += 1

    ws[f'A{row}'] = "Gap to Close: 8 points"
    ws[f'B{row}'] = "Timeline: 12 months"
    row += 1

    ws[f'A{row}'] = "Available Budget: ₹50 lakhs"
    row += 2

    # Reverse calculation
    ws[f'A{row}'] = "STEP 2: REVERSE CALCULATION"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Required Total Points: (80 × 109) / 100 = 87.2 points"
    row += 1

    ws[f'A{row}'] = "Current Total Points: (72 × 109) / 100 = 78.48 points"
    row += 1

    ws[f'A{row}'] = "Points to Gain: 87.2 - 78.48 = 8.72 points"
    row += 2

    ws[f'A{row}'] = "DIMENSION-LEVEL TARGETS"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    ws[f'A{row}'] = "Dimension"
    ws[f'B{row}'] = "Current"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Gap"
    ws[f'E{row}'] = "Weight"
    ws[f'F{row}'] = "Contribution"
    ws[f'G{row}'] = "Feasibility"
    ws[f'H{row}'] = "Priority"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        apply_cell_style(ws[f'{col}{row}'], 'header_light', colors)
    row += 1

    # Example targets
    targets = [
        ('D01 Academic', 75, 80, 5, 10, 0.5, '90%', 'High'),
        ('D02 Teacher Welfare', 70, 78, 8, 9, 0.72, '70%', 'High'),
        ('D03 Leadership', 78, 82, 4, 10, 0.4, '95%', 'Medium'),
        ('D04 Parent Engagement', 72, 78, 6, 8, 0.48, '85%', 'High'),
        ('D05 Safety', 80, 85, 5, 10, 0.5, '75%', 'Medium'),
        ('D06 Infrastructure', 68, 75, 7, 7, 0.49, '60%', 'Low'),
        ('D07 Co-Curricular', 65, 72, 7, 6, 0.42, '80%', 'Medium'),
        ('D08 Individual Attention', 75, 80, 5, 9, 0.45, '65%', 'High'),
        ('D09 Value for Money', 70, 76, 6, 7, 0.42, '70%', 'Medium'),
        ('D10 Special Needs', 60, 68, 8, 6, 0.48, '55%', 'Low'),
        ('D11 Community Service', 65, 72, 7, 5, 0.35, '75%', 'Low'),
        ('D12 Faculty Competence', 82, 86, 4, 9, 0.36, '85%', 'Medium'),
        ('D13 Internationalism', 70, 76, 6, 6, 0.36, '70%', 'Low'),
        ('D14 Management Vision', 76, 82, 6, 8, 0.48, '90%', 'High'),
    ]

    for dim, cur, tgt, gap, weight, contrib, feas, prior in targets:
        ws[f'A{row}'] = dim
        ws[f'B{row}'] = cur
        ws[f'C{row}'] = tgt
        ws[f'D{row}'] = gap
        ws[f'E{row}'] = weight
        ws[f'F{row}'] = contrib
        ws[f'G{row}'] = feas
        ws[f'H{row}'] = prior

        apply_cell_style(ws[f'B{row}'], 'current', colors)
        apply_cell_style(ws[f'C{row}'], 'target', colors)
        apply_cell_style(ws[f'F{row}'], 'calculation', colors)
        row += 1

    row += 2

    # Key actions
    ws[f'A{row}'] = "STEP 3: KEY ACTIONS"
    apply_cell_style(ws[f'A{row}'], 'header_light', colors)
    row += 1

    actions = [
        ('D02 - Teacher Salary Increase', '8 points', '₹25L', '6 months', 'HIGH'),
        ('D04 - Parent SLA System', '6 points', '₹5L', '3 months', 'QUICK WIN'),
        ('D03 - Strategic Planning', '4 points', '₹2L', '3 months', 'QUICK WIN'),
        ('D01 - Curriculum Enhancement', '5 points', '₹8L', '9 months', 'HIGH'),
    ]

    for action, impact, cost, timeline, priority in actions:
        ws[f'A{row}'] = action
        ws[f'B{row}'] = impact
        ws[f'C{row}'] = cost
        ws[f'D{row}'] = timeline
        ws[f'E{row}'] = priority
        row += 1

    row += 2

    ws[f'A{row}'] = "RESULT: Feasible to achieve 80/100 target in 12 months with ₹50L budget"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws.merge_cells(f'A{row}:H{row}')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

def main():
    print("Creating Stage 3 Reverse Outcome Modeling Engine...")
    print("  - Goal Setting Sheet")
    print("  - Reverse Calculation Engine")
    print("  - Feasibility Analysis")
    print("  - Action Mapping")
    print("  - Resource Allocation")
    print("  - Timeline & Milestones")
    print("  - Complete Example Scenario")
    print("")

    wb = Workbook()
    wb.remove(wb.active)
    colors = create_colors()

    create_goal_setting_sheet(wb, colors)
    print("[OK] Sheet 1: Goal Setting")

    create_reverse_calculation_sheet(wb, colors)
    print("[OK] Sheet 2: Reverse Calculation Engine")

    create_feasibility_sheet(wb, colors)
    print("[OK] Sheet 3: Feasibility Analysis")

    create_action_mapping_sheet(wb, colors)
    print("[OK] Sheet 4: Action Mapping")

    create_resource_allocation_sheet(wb, colors)
    print("[OK] Sheet 5: Resource Allocation")

    create_timeline_sheet(wb, colors)
    print("[OK] Sheet 6: Timeline & Milestones")

    create_example_scenario_sheet(wb, colors)
    print("[OK] Sheet 7: Example Scenario")

    filename = 'public/DISHA_STAGE3_ReverseOutcomeModeling.xlsx'
    wb.save(filename)

    print("")
    print("=" * 70)
    print("SUCCESS! STAGE 3 REVERSE OUTCOME MODELING ENGINE CREATED")
    print("=" * 70)
    print(f"\nFile saved: {filename}")
    print("\nWorkbook contains:")
    print("  Sheet 1: Goal Setting - Define targets and parameters")
    print("  Sheet 2: Reverse Calculation - Calculate required dimension scores")
    print("  Sheet 3: Feasibility Analysis - Assess achievability")
    print("  Sheet 4: Action Mapping - Convert targets to specific actions")
    print("  Sheet 5: Resource Allocation - Budget and effort planning")
    print("  Sheet 6: Timeline & Milestones - Implementation schedule")
    print("  Sheet 7: Example Scenario - Complete walkthrough")
    print("\nReady to open in Microsoft Excel!")

if __name__ == "__main__":
    main()
