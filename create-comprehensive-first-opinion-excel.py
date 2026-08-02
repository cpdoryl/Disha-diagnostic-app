#!/usr/bin/env python3
"""
Create comprehensive First Opinion Engine Excel workbook with ALL 15 challenges,
multiple example scenarios, and complete measurement engine walkthrough
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_colors():
    """Define color scheme"""
    return {
        'header_dark': '1E40AF',      # Dark blue
        'header_light': '0D9488',     # Teal
        'input_field': 'CCFBF1',      # Light cyan
        'calculation': 'FEF3C7',      # Yellow
        'result': 'FFFFE4E6',         # Light red/pink
        'success': 'D1FAE5',          # Light green
        'warning': 'FED7AA',          # Orange
        'danger': 'FEE2E2',           # Red
        'neutral': 'F3F4F6'           # Gray
    }

# Complete questionnaire data for all 15 challenges
COMPLETE_QUESTIONNAIRES = {
    'C1': {
        'name': 'Enrollment Decline',
        'category': 'Growth & Enrollment',
        'metric': 'New Student Intake Rate (%), Student Retention Rate (%)',
        'questions': [
            {
                'q': 'Q1.1: What is the trend of new enrollments in the last 3 years?',
                'options': [
                    ('Strong growth (>20% YoY)', 1),
                    ('Moderate growth (10-20% YoY)', 2),
                    ('Flat/minimal growth (<10% YoY)', 4),
                    ('Slight decline (-5% to -10%)', 6),
                    ('Moderate decline (-10% to -20%)', 8),
                    ('Severe decline (<-20%)', 10)
                ]
            },
            {
                'q': 'Q1.2: How competitive is enrollment vs peer schools in your market?',
                'options': [
                    ('More competitive than peers', 1),
                    ('Equally competitive', 3),
                    ('Less competitive than peers', 5),
                    ('Significantly less competitive', 8),
                    ('Major disadvantage vs competitors', 10)
                ]
            },
            {
                'q': 'Q1.3: What is your student retention rate from Grade 1 to Grade 12?',
                'options': [
                    ('90%+ retention', 1),
                    ('80-90% retention', 3),
                    ('70-80% retention', 5),
                    ('60-70% retention', 7),
                    ('<60% retention', 10)
                ]
            }
        ]
    },
    'C2': {
        'name': 'Student Attrition',
        'category': 'Growth & Enrollment',
        'metric': 'Mid-Year Dropout Rate (%), Outflow to Competitors (%)',
        'questions': [
            {
                'q': 'Q2.1: What percentage of students leave mid-year?',
                'options': [
                    ('0-2% mid-year attrition', 1),
                    ('2-5% mid-year attrition', 3),
                    ('5-8% mid-year attrition', 5),
                    ('8-12% mid-year attrition', 7),
                    ('>12% mid-year attrition', 10)
                ]
            },
            {
                'q': 'Q2.2: What are the primary reasons for student exit?',
                'options': [
                    ('Academic/personal - isolated cases', 2),
                    ('Mixed reasons, mostly controllable', 4),
                    ('Mostly controllable (affordability, safety)', 6),
                    ('Major controllable issues (staff, quality)', 8),
                    ('Systemic failure in core offering', 10)
                ]
            },
            {
                'q': 'Q2.3: How many students shift to competitor schools annually?',
                'options': [
                    ('Very few (<2%)', 1),
                    ('Some (2-5%)', 3),
                    ('Noticeable (5-10%)', 5),
                    ('Significant (10-15%)', 8),
                    ('Severe (>15%)', 10)
                ]
            }
        ]
    },
    'C3': {
        'name': 'Fee Collection Challenges',
        'category': 'Growth & Enrollment',
        'metric': 'Fee Realization Rate (%), Days Sales Outstanding (DSO)',
        'questions': [
            {
                'q': 'Q3.1: What percentage of annual fees is realized?',
                'options': [
                    ('95-100% realization', 1),
                    ('90-95% realization', 2),
                    ('85-90% realization', 4),
                    ('75-85% realization', 7),
                    ('<75% realization', 10)
                ]
            },
            {
                'q': 'Q3.2: What is the average payment delay from parents?',
                'options': [
                    ('On-time or early payment', 1),
                    ('30 days average delay', 2),
                    ('60 days average delay', 4),
                    ('90+ days average delay', 7),
                    ('Chronic defaults and disputes', 10)
                ]
            },
            {
                'q': 'Q3.3: How many parents request scholarships/fee reduction?',
                'options': [
                    ('<5% requesting concession', 1),
                    ('5-10% requesting concession', 2),
                    ('10-20% requesting concession', 4),
                    ('20-30% requesting concession', 7),
                    ('>30% requesting concession', 10)
                ]
            }
        ]
    },
    'C4': {
        'name': 'Teacher Attrition',
        'category': 'People & Staffing',
        'metric': 'Teacher Turnover Rate (%), Avg Teacher Tenure (years)',
        'questions': [
            {
                'q': 'Q4.1: What is your annual teacher attrition rate?',
                'options': [
                    ('<5% annual turnover', 1),
                    ('5-10% annual turnover', 2),
                    ('10-15% annual turnover', 4),
                    ('15-25% annual turnover', 7),
                    ('>25% annual turnover', 10)
                ]
            },
            {
                'q': 'Q4.2: What is the primary reason teachers leave?',
                'options': [
                    ('Retirement/personal - unavoidable', 1),
                    ('Career growth opportunities elsewhere', 3),
                    ('Better compensation packages', 5),
                    ('Management/culture issues', 7),
                    ('Systemic institutional failure', 10)
                ]
            },
            {
                'q': 'Q4.3: What is the average teacher tenure at your school?',
                'options': [
                    ('10+ years average', 1),
                    ('7-10 years average', 2),
                    ('5-7 years average', 4),
                    ('3-5 years average', 7),
                    ('<3 years average', 10)
                ]
            }
        ]
    },
    'C5': {
        'name': 'Staff Capability Gaps',
        'category': 'People & Staffing',
        'metric': 'Teacher Competency Score (%), Professional Qualification %',
        'questions': [
            {
                'q': 'Q5.1: What % of your teachers have subject specialist qualifications?',
                'options': [
                    ('>90% specialists', 1),
                    ('80-90% specialists', 2),
                    ('70-80% specialists', 4),
                    ('50-70% specialists', 6),
                    ('<50% specialists', 10)
                ]
            },
            {
                'q': 'Q5.2: How aligned is teacher competency to curriculum needs?',
                'options': [
                    ('Highly aligned - excellent match', 1),
                    ('Well aligned - good match', 2),
                    ('Mostly aligned - acceptable', 4),
                    ('Gaps in key subjects/areas', 7),
                    ('Major misalignment issues', 10)
                ]
            },
            {
                'q': 'Q5.3: How often do teachers receive upskilling training?',
                'options': [
                    ('Regular (monthly+)', 1),
                    ('Quarterly', 2),
                    ('Semi-annual', 3),
                    ('Annual', 5),
                    ('Rarely/never', 10)
                ]
            }
        ]
    },
    'C6': {
        'name': 'Leadership Capability Gap',
        'category': 'People & Staffing',
        'metric': 'Leadership Competency Score (%), Principal/VP Experience (years)',
        'questions': [
            {
                'q': 'Q6.1: How many years of leadership experience does your top leader have?',
                'options': [
                    ('15+ years', 1),
                    ('10-15 years', 2),
                    ('7-10 years', 3),
                    ('3-7 years', 5),
                    ('<3 years', 10)
                ]
            },
            {
                'q': 'Q6.2: How clear is the institutional vision and strategy?',
                'options': [
                    ('Crystal clear, well communicated', 1),
                    ('Clear, mostly aligned', 2),
                    ('Somewhat clear, partial alignment', 4),
                    ('Unclear, limited buy-in', 7),
                    ('No clear vision/direction', 10)
                ]
            },
            {
                'q': 'Q6.3: How effective is decision-making and governance?',
                'options': [
                    ('Excellent - quick, well-informed', 1),
                    ('Good - timely, mostly sound', 2),
                    ('Adequate - but some delays', 4),
                    ('Poor - slow, inconsistent', 7),
                    ('Dysfunctional - unclear authority', 10)
                ]
            }
        ]
    },
    'C7': {
        'name': 'Academic Quality Decline',
        'category': 'Academic & Wellbeing',
        'metric': 'Board Exam Pass Rate (%), Average Subject Score (%)',
        'questions': [
            {
                'q': 'Q7.1: What is your board exam pass rate?',
                'options': [
                    ('>95% pass rate', 1),
                    ('90-95% pass rate', 2),
                    ('85-90% pass rate', 3),
                    ('70-85% pass rate', 5),
                    ('<70% pass rate', 10)
                ]
            },
            {
                'q': 'Q7.2: How are your scores trending vs peer institutions?',
                'options': [
                    ('Consistently above peer average', 1),
                    ('At or near peer average', 2),
                    ('Slightly below peer average', 4),
                    ('Significantly below peer average', 7),
                    ('Much lower than peers', 10)
                ]
            },
            {
                'q': 'Q7.3: What % of students score above 70% aggregate?',
                'options': [
                    ('>80% high achievers', 1),
                    ('70-80% high achievers', 2),
                    ('50-70% high achievers', 4),
                    ('30-50% high achievers', 7),
                    ('<30% high achievers', 10)
                ]
            }
        ]
    },
    'C8': {
        'name': 'Student Wellbeing Issues',
        'category': 'Academic & Wellbeing',
        'metric': 'Mental Health Incidents (per 1000), Safety Violations (count/year)',
        'questions': [
            {
                'q': 'Q8.1: How many mental health/psychological issues are reported?',
                'options': [
                    ('Very few - excellent support', 1),
                    ('Minimal - good awareness', 2),
                    ('Some - moderate support', 4),
                    ('Multiple - limited support', 7),
                    ('Significant - systemic issues', 10)
                ]
            },
            {
                'q': 'Q8.2: How safe do students feel at school?',
                'options': [
                    ('Extremely safe - zero incidents', 1),
                    ('Very safe - rare incidents', 2),
                    ('Safe - occasional issues', 4),
                    ('Somewhat unsafe - regular issues', 7),
                    ('Unsafe - frequent incidents', 10)
                ]
            },
            {
                'q': 'Q8.3: How strong is peer bullying/harassment prevention?',
                'options': [
                    ('Excellent - proactive culture', 1),
                    ('Good - reported and addressed', 2),
                    ('Adequate - some gaps', 4),
                    ('Poor - incidents go unaddressed', 7),
                    ('Severe - systemic bullying', 10)
                ]
            }
        ]
    },
    'C9': {
        'name': 'Remedial Lag',
        'category': 'Academic & Wellbeing',
        'metric': 'Remedial Support Coverage (%), Improvement Rate (%)',
        'questions': [
            {
                'q': 'Q9.1: What % of students requiring remedial support receive it?',
                'options': [
                    ('90%+ support provided', 1),
                    ('75-90% support provided', 2),
                    ('50-75% support provided', 4),
                    ('25-50% support provided', 7),
                    ('<25% support provided', 10)
                ]
            },
            {
                'q': 'Q9.2: How effective is remedial intervention (improvement rate)?',
                'options': [
                    ('Very effective - 70%+ improve', 1),
                    ('Effective - 50-70% improve', 2),
                    ('Moderate - 30-50% improve', 4),
                    ('Limited - 10-30% improve', 7),
                    ('Ineffective - <10% improve', 10)
                ]
            },
            {
                'q': 'Q9.3: How many grade levels receive remedial support?',
                'options': [
                    ('All grades, all subjects', 1),
                    ('Most grades, key subjects', 2),
                    ('Some grades, core subjects', 4),
                    ('Limited grades only', 7),
                    ('No systematic remedial', 10)
                ]
            }
        ]
    },
    'C10': {
        'name': 'Parent Communication Issues',
        'category': 'Reputation & Competition',
        'metric': 'Parent Satisfaction Score (%), Parent Response Rate (%)',
        'questions': [
            {
                'q': 'Q10.1: How satisfied are parents with communication?',
                'options': [
                    ('Very satisfied - 90%+ positive', 1),
                    ('Satisfied - 75-90% positive', 2),
                    ('Neutral - 50-75% positive', 4),
                    ('Dissatisfied - 25-50% positive', 7),
                    ('Very dissatisfied - <25% positive', 10)
                ]
            },
            {
                'q': 'Q10.2: What is your average response time to parent queries?',
                'options': [
                    ('Within 4 hours', 1),
                    ('Within 12 hours', 2),
                    ('Within 24 hours', 4),
                    ('2-3 days', 7),
                    ('>3 days / no response', 10)
                ]
            }
        ]
    },
    'C11': {
        'name': 'Competitive Pressure',
        'category': 'Reputation & Competition',
        'metric': 'Market Share Loss (%), Competitor Win Rate (%)',
        'questions': [
            {
                'q': 'Q11.1: How intense is competition in your market?',
                'options': [
                    ('Limited - clear market leader', 1),
                    ('Moderate - few strong competitors', 2),
                    ('High - several strong competitors', 4),
                    ('Very high - many aggressive competitors', 7),
                    ('Extreme - commoditized market', 10)
                ]
            },
            {
                'q': 'Q11.2: Are you losing enrollment to specific competitors?',
                'options': [
                    ('Not losing market share', 1),
                    ('Minimal losses (<2% annually)', 2),
                    ('Noticeable losses (2-5% annually)', 4),
                    ('Significant losses (5-10% annually)', 7),
                    ('Severe losses (>10% annually)', 10)
                ]
            }
        ]
    },
    'C12': {
        'name': 'Brand/Reputation Issues',
        'category': 'Reputation & Competition',
        'metric': 'Brand Perception Score (%), Media Sentiment (%)',
        'questions': [
            {
                'q': 'Q12.1: How is your school perceived by target parents?',
                'options': [
                    ('Excellent reputation - top of mind', 1),
                    ('Good reputation - well regarded', 2),
                    ('Neutral reputation - known entity', 4),
                    ('Poor reputation - perception concerns', 7),
                    ('Very poor reputation - negative image', 10)
                ]
            },
            {
                'q': 'Q12.2: How often does negative news/media coverage appear?',
                'options': [
                    ('Never - only positive coverage', 1),
                    ('Rare - occasional positive stories', 2),
                    ('Occasional - mixed coverage', 4),
                    ('Regular - notable negative stories', 7),
                    ('Frequent - serious reputation damage', 10)
                ]
            }
        ]
    },
    'C13': {
        'name': 'Cost Inflation',
        'category': 'Operations & Finance',
        'metric': 'Cost Increase YoY (%), Operating Margin (%)',
        'questions': [
            {
                'q': 'Q13.1: What is your cost inflation rate vs fee increase?',
                'options': [
                    ('Inflation lower than fee increase', 1),
                    ('Aligned with fee increase', 2),
                    ('Slightly higher than fee increase', 4),
                    ('Significantly higher than fee increase', 7),
                    ('Costs growing much faster than fees', 10)
                ]
            },
            {
                'q': 'Q13.2: How healthy is your operating margin?',
                'options': [
                    ('Excellent - 20%+ margin', 1),
                    ('Good - 15-20% margin', 2),
                    ('Adequate - 10-15% margin', 4),
                    ('Thin - 5-10% margin', 7),
                    ('Negative - losses or break-even', 10)
                ]
            }
        ]
    },
    'C14': {
        'name': 'Infrastructure Deficits',
        'category': 'Operations & Finance',
        'metric': 'Infrastructure Quality Score (%), Maintenance Backlog (₹)',
        'questions': [
            {
                'q': 'Q14.1: How would you rate your overall infrastructure quality?',
                'options': [
                    ('Excellent - modern, well-maintained', 1),
                    ('Good - mostly adequate, minor updates needed', 2),
                    ('Fair - functional but aging', 4),
                    ('Poor - significant deficits', 7),
                    ('Very poor - major infrastructure issues', 10)
                ]
            },
            {
                'q': 'Q14.2: What is your infrastructure maintenance backlog?',
                'options': [
                    ('Current - all maintained', 1),
                    ('Minimal - minor backlog', 2),
                    ('Moderate - significant backlog', 4),
                    ('Large - major deferred maintenance', 7),
                    ('Severe - critical backlog', 10)
                ]
            }
        ]
    },
    'C15': {
        'name': 'Compliance & Regulatory Stress',
        'category': 'Operations & Finance',
        'metric': 'Compliance Score (%), Regulatory Violations (count/year)',
        'questions': [
            {
                'q': 'Q15.1: How well are you complying with education regulations?',
                'options': [
                    ('Full compliance - zero violations', 1),
                    ('Nearly compliant - minor issues', 2),
                    ('Mostly compliant - some gaps', 4),
                    ('Non-compliant - multiple violations', 7),
                    ('Severe violations - regulatory risk', 10)
                ]
            },
            {
                'q': 'Q15.2: Are there pending compliance audits or notices?',
                'options': [
                    ('No pending issues', 1),
                    ('Routine audit cycle', 2),
                    ('Minor notices to address', 4),
                    ('Significant compliance notices', 7),
                    ('Serious regulatory action pending', 10)
                ]
            }
        ]
    }
}

def apply_cell_style(cell, style_type, colors):
    """Apply predefined styles to cells"""
    if style_type == 'header_dark':
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=colors['header_dark'], end_color=colors['header_dark'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    elif style_type == 'header_light':
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=colors['header_light'], end_color=colors['header_light'], fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    elif style_type == 'input':
        cell.fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif style_type == 'calculation':
        cell.fill = PatternFill(start_color=colors['calculation'], end_color=colors['calculation'], fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif style_type == 'result':
        cell.fill = PatternFill(start_color=colors['result'], end_color=colors['result'], fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def create_all_challenges_questionnaire(wb, colors):
    """Create sheet with ALL 15 complete questionnaires"""
    ws = wb.create_sheet("ALL 15 CHALLENGES QUESTIONNAIRE")

    row = 1
    ws[f'A{row}'] = "COMPLETE SCREENING QUESTIONNAIRE - ALL 15 CHALLENGES"
    apply_cell_style(ws[f'A{row}'], 'header_dark', colors)
    ws.merge_cells(f'A{row}:F{row}')
    row += 2

    for challenge_id in ['C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9', 'C10', 'C11', 'C12', 'C13', 'C14', 'C15']:
        challenge = COMPLETE_QUESTIONNAIRES[challenge_id]

        # Challenge header
        ws[f'A{row}'] = f"{challenge_id}: {challenge['name']}"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF", size=10)
        ws[f'A{row}'].fill = PatternFill(start_color=colors['header_light'], end_color=colors['header_light'], fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        ws[f'A{row}'] = f"Category: {challenge['category']} | Metric: {challenge['metric']}"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # Questions for this challenge
        for q_idx, question in enumerate(challenge['questions'], 1):
            ws[f'A{row}'] = question['q']
            ws[f'A{row}'].font = Font(bold=True, size=9)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Options with weights
            for option_text, weight in question['options']:
                ws[f'A{row}'] = f"  {option_text}"
                ws[f'B{row}'] = f"Weight: {weight}"
                ws[f'B{row}'].fill = PatternFill(start_color=colors['input_field'], end_color=colors['input_field'], fill_type="solid")
                row += 1

            row += 1  # Spacing

        row += 1  # Challenge spacing

    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

def create_multiple_scenarios(wb, colors):
    """Create sheet with MULTIPLE example scenarios and calculations"""
    ws = wb.create_sheet("SCENARIO EXAMPLES & CALCULATIONS")

    scenarios = [
        {
            'name': 'Scenario 1: Enrollment Crisis Focus',
            'school': 'Delhi Public Academy',
            'challenges': ['C1', 'C2', 'C3'],
            'weights': [0.50, 0.30, 0.20],
            'answers': {
                'C1': [6, 7, 8],  # Slight decline, less competitive, 60-70% retention
                'C2': [5, 6, 6],  # 5-8% attrition, mixed reasons, 5-10% to competitors
                'C3': [4, 3, 4]   # 85-90% realization, 30d delay, 10-20% requesting
            },
            'metrics': {'STR': 28, 'SLA': 24, 'Training': 15, 'Planning': 4}
        },
        {
            'name': 'Scenario 2: Quality & Academic Focus',
            'school': 'Mumbai High Excellence',
            'challenges': ['C7', 'C8', 'C9', 'C4'],
            'weights': [0.35, 0.25, 0.25, 0.15],
            'answers': {
                'C7': [4, 5, 5],  # 85-90% pass, below peer average, 50-70% high achievers
                'C8': [4, 5, 5],  # Some issues, somewhat unsafe, some bullying
                'C9': [5, 4, 5],  # 25-50% support, limited improvement, limited grades
                'C4': [3, 2, 3]   # 5-10% turnover, career opportunities, 7-10yr tenure
            },
            'metrics': {'STR': 32, 'SLA': 36, 'Training': 18, 'Planning': 3}
        },
        {
            'name': 'Scenario 3: Multi-Domain Challenge',
            'school': 'Bangalore Growth Institute',
            'challenges': ['C1', 'C4', 'C10', 'C12', 'C13'],
            'weights': [0.25, 0.20, 0.20, 0.20, 0.15],
            'answers': {
                'C1': [4, 5, 6],
                'C4': [4, 3, 4],
                'C10': [5, 6],
                'C12': [5, 6],
                'C13': [4, 5]
            },
            'metrics': {'STR': 35, 'SLA': 42, 'Training': 12, 'Planning': 2}
        }
    ]

    row = 1

    for scenario_idx, scenario in enumerate(scenarios):
        # Scenario header
        ws[f'A{row}'] = scenario['name']
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF", size=12)
        ws[f'A{row}'].fill = PatternFill(start_color=colors['header_dark'], end_color=colors['header_dark'], fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # School info
        ws[f'A{row}'] = f"School: {scenario['school']}"
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # Challenge selection
        ws[f'A{row}'] = "Selected Challenges:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        for ch_id, weight in zip(scenario['challenges'], scenario['weights']):
            ch_name = COMPLETE_QUESTIONNAIRES[ch_id]['name']
            ws[f'A{row}'] = f"  {ch_id}: {ch_name} - Weight: {weight*100:.0f}%"
            row += 1

        row += 1

        # Screening answers
        ws[f'A{row}'] = "SCREENING ANSWERS & CALCULATION"
        apply_cell_style(ws[f'A{row}'], 'header_light', colors)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        total_weight = 0
        challenge_contributions = {}

        for ch_id in scenario['challenges']:
            ch_name = COMPLETE_QUESTIONNAIRES[ch_id]['name']
            answers = scenario['answers'][ch_id]
            weight = scenario['weights'][scenario['challenges'].index(ch_id)]

            ws[f'A{row}'] = ch_name
            ws[f'A{row}'].font = Font(bold=True, size=9)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Show questions and selected weights
            for q_idx, (q_data, answer_weight) in enumerate(zip(COMPLETE_QUESTIONNAIRES[ch_id]['questions'], answers)):
                q_text = f"Q: {q_data['q'][:60]}..."
                ws[f'A{row}'] = q_text
                ws[f'B{row}'] = f"Weight: {answer_weight}/10"
                apply_cell_style(ws[f'B{row}'], 'input', colors)
                row += 1

            # Challenge contribution to S_sub
            challenge_sum = sum(answers)
            max_possible = len(answers) * 10
            challenge_contribution = (challenge_sum / max_possible) * weight * 100
            challenge_contributions[ch_id] = challenge_contribution
            total_weight += challenge_sum * weight

            ws[f'A{row}'] = f"  {ch_name} Sum: {challenge_sum}/{ len(answers)*10} × Weight {weight*100:.0f}%"
            ws[f'B{row}'] = f"{challenge_contribution:.1f}%"
            apply_cell_style(ws[f'B{row}'], 'calculation', colors)
            row += 2

        row += 1

        # S_sub calculation
        ws[f'A{row}'] = "S_sub (Subjective Base Score) Calculation:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        max_possible_all = sum(len(COMPLETE_QUESTIONNAIRES[ch_id]['questions']) * 10 for ch_id in scenario['challenges'])
        actual_sum = sum(sum(scenario['answers'][ch_id]) for ch_id in scenario['challenges'])
        percentage_of_max = (actual_sum / max_possible_all) * 100
        s_sub = 100 - percentage_of_max

        ws[f'A{row}'] = f"Total Response Weight: {actual_sum}/{max_possible_all}"
        ws[f'B{row}'] = f"{percentage_of_max:.1f}%"
        apply_cell_style(ws[f'B{row}'], 'input', colors)
        row += 1

        ws[f'A{row}'] = f"S_sub = 100 - {percentage_of_max:.1f}%"
        ws[f'B{row}'] = f"{s_sub:.1f}"
        apply_cell_style(ws[f'B{row}'], 'calculation', colors)
        row += 2

        # Operational metrics
        ws[f'A{row}'] = "OPERATIONAL METRICS"
        apply_cell_style(ws[f'A{row}'], 'header_light', colors)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        metrics = scenario['metrics']
        m_str = 1.05 if metrics['STR'] <= 20 else (1.00 if metrics['STR'] <= 28 else (0.88 if metrics['STR'] <= 35 else 0.75))
        m_sla = 1.00 if metrics['SLA'] <= 12 else (0.95 if metrics['SLA'] <= 24 else (0.70 if metrics['SLA'] <= 48 else 0.50))
        m_train = 1.00 if metrics['Training'] >= 25 else (0.85 if metrics['Training'] >= 15 else 0.60)
        m_plan = 1.00 if metrics['Planning'] >= 5 else (0.88 if metrics['Planning'] >= 3 else 0.75)

        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Benchmark"
        ws[f'D{row}'] = "Multiplier"
        for col in ['A', 'B', 'C', 'D']:
            apply_cell_style(ws[f'{col}{row}'], 'header_light', colors)
        row += 1

        ws[f'A{row}'] = "Student-Teacher Ratio"
        ws[f'B{row}'] = metrics['STR']
        apply_cell_style(ws[f'B{row}'], 'input', colors)
        ws[f'C{row}'] = "≤20"
        ws[f'D{row}'] = f"{m_str:.2f}"
        apply_cell_style(ws[f'D{row}'], 'calculation', colors)
        row += 1

        ws[f'A{row}'] = "Parent Response SLA (hours)"
        ws[f'B{row}'] = metrics['SLA']
        apply_cell_style(ws[f'B{row}'], 'input', colors)
        ws[f'C{row}'] = "≤12"
        ws[f'D{row}'] = f"{m_sla:.2f}"
        apply_cell_style(ws[f'D{row}'], 'calculation', colors)
        row += 1

        ws[f'A{row}'] = "Teacher Training (hrs/year)"
        ws[f'B{row}'] = metrics['Training']
        apply_cell_style(ws[f'B{row}'], 'input', colors)
        ws[f'C{row}'] = "≥25"
        ws[f'D{row}'] = f"{m_train:.2f}"
        apply_cell_style(ws[f'D{row}'], 'calculation', colors)
        row += 1

        ws[f'A{row}'] = "Weekly Planning Time (hrs)"
        ws[f'B{row}'] = metrics['Planning']
        apply_cell_style(ws[f'B{row}'], 'input', colors)
        ws[f'C{row}'] = "≥5"
        ws[f'D{row}'] = f"{m_plan:.2f}"
        apply_cell_style(ws[f'D{row}'], 'calculation', colors)
        row += 2

        # M_obj calculation
        m_obj = m_str * m_sla * m_train * m_plan
        ws[f'A{row}'] = "M_obj (Objective Multiplier):"
        ws[f'B{row}'] = f"{m_str:.2f} × {m_sla:.2f} × {m_train:.2f} × {m_plan:.2f}"
        ws[f'C{row}'] = f"{m_obj:.4f}"
        apply_cell_style(ws[f'C{row}'], 'calculation', colors)
        row += 2

        # Final calculation
        scaled_score = s_sub * m_obj
        delusion_penalty = 0 if s_sub < 80 else (s_sub - 80)
        health_index = max(0, min(100, scaled_score - delusion_penalty))

        ws[f'A{row}'] = "FINAL HEALTH INDEX CALCULATION"
        apply_cell_style(ws[f'A{row}'], 'header_light', colors)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = f"Scaled Score: {s_sub:.1f} × {m_obj:.4f}"
        ws[f'B{row}'] = f"{scaled_score:.2f}"
        apply_cell_style(ws[f'B{row}'], 'calculation', colors)
        row += 1

        ws[f'A{row}'] = f"Delusion Penalty (if S_sub ≥ 80): {delusion_penalty:.2f}"
        ws[f'B{row}'] = "0" if delusion_penalty == 0 else str(delusion_penalty)
        apply_cell_style(ws[f'B{row}'], 'calculation', colors)
        row += 1

        ws[f'A{row}'] = "HEALTH INDEX (H):"
        ws[f'B{row}'] = f"{health_index:.2f}/100"
        apply_cell_style(ws[f'B{row}'], 'result', colors)
        row += 1

        # Risk classification
        if health_index >= 70:
            risk = "ELITE EQUILIBRIUM (Green - Excellent)"
        elif health_index >= 50:
            risk = "HIDDEN EXCELLENCE (Yellow - Good)"
        elif health_index >= 30:
            risk = "DELUSIONAL COMFORT (Orange - At Risk)"
        else:
            risk = "CRITICAL COLLAPSE (Red - High Risk)"

        ws[f'A{row}'] = "RISK CLASSIFICATION:"
        ws[f'B{row}'] = risk
        apply_cell_style(ws[f'B{row}'], 'result', colors)
        row += 3

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

def create_measurement_engine(wb, colors):
    """Create sheet explaining the measurement engine in detail"""
    ws = wb.create_sheet("MEASUREMENT ENGINE")

    row = 1
    ws[f'A{row}'] = "DISHA FIRST OPINION ENGINE - MEASUREMENT METHODOLOGY"
    apply_cell_style(ws[f'A{row}'], 'header_dark', colors)
    ws.merge_cells(f'A{row}:E{row}')
    row += 2

    sections = [
        {
            'title': '1. CHALLENGE SELECTION & WEIGHTING',
            'content': [
                'Schools select 1-5 challenges from 15 available categories',
                'Each selected challenge is assigned a weight (% importance)',
                'Weights must sum to 100%',
                'Weights reflect which challenges have highest impact on school'
            ]
        },
        {
            'title': '2. SCREENING QUESTIONNAIRE RESPONSE',
            'content': [
                'For each challenge, 2-3 screening questions are asked',
                'Each question has 5-6 ordinal options with weights 1-10',
                'Weight 1 = Excellent/No issue',
                'Weight 10 = Severe/Critical issue',
                'Parent/staff select most applicable option',
                'System records the numeric weight (1-10) for that answer'
            ]
        },
        {
            'title': '3. SUBJECTIVE SCORE (S_sub) CALCULATION',
            'content': [
                'For each challenge: Sum all question weights',
                'Calculate percentage: Response Score / Maximum Possible',
                'Maximum possible = Number of Questions × 10',
                'S_sub = 100 - (Average %-age across all questions)',
                'S_sub ranges from 0-100 (100=healthy, 0=critical)',
                'Example: 3 questions with weights [6,7,8], max=30',
                '  Sum=21, Percentage=21/30=70%',
                '  S_sub = 100-70 = 30'
            ]
        },
        {
            'title': '4. OBJECTIVE MULTIPLIERS (M_obj) CALCULATION',
            'content': [
                'School provides 4 operational metrics:',
                '  • Student-Teacher Ratio (STR)',
                '  • Parent Response SLA (hours)',
                '  • Annual Teacher Training (hours)',
                '  • Weekly Planning Time (hours)',
                '',
                'STR Multiplier (m_STR):',
                '  ≤20 students/teacher = 1.05 (excellent)',
                '  21-28 students/teacher = 1.00 (good)',
                '  29-35 students/teacher = 0.88 (acceptable)',
                '  >35 students/teacher = 0.75 (poor)',
                '',
                'Parent SLA Multiplier (m_SLA):',
                '  ≤12 hours = 1.00 (excellent)',
                '  13-24 hours = 0.95 (good)',
                '  25-48 hours = 0.70 (acceptable)',
                '  >48 hours = 0.50 (poor)',
                '',
                'Training Multiplier (m_train):',
                '  ≥25 hours/year = 1.00 (excellent)',
                '  15-24 hours/year = 0.85 (acceptable)',
                '  <15 hours/year = 0.60 (poor)',
                '',
                'Planning Time Multiplier (m_plan):',
                '  ≥5 hours/week = 1.00 (excellent)',
                '  3-5 hours/week = 0.88 (acceptable)',
                '  <3 hours/week = 0.75 (poor)',
                '',
                'M_obj = m_STR × m_SLA × m_train × m_plan',
                'M_obj ranges from 0.21 to 1.05'
            ]
        },
        {
            'title': '5. SCALED HEALTH SCORE CALCULATION',
            'content': [
                'Scaled Score = S_sub × M_obj',
                'This applies objective reality to subjective perception',
                'If leadership overestimates health (high S_sub) but metrics',
                'are poor (low M_obj), scaled score reflects reality',
                'Range: 0-100'
            ]
        },
        {
            'title': '6. DELUSION PENALTY',
            'content': [
                'Identifies when leadership is overconfident',
                'If S_sub ≥ 80 AND M_obj < 0.7:',
                '  Delusion Penalty = S_sub - 80',
                'This penalty is subtracted from final score',
                'If S_sub < 80: No penalty (realistic self-assessment)'
            ]
        },
        {
            'title': '7. FINAL HEALTH INDEX (H)',
            'content': [
                'H = MAX(0, MIN(100, Scaled Score - Delusion Penalty))',
                'Ensures score stays within 0-100 range',
                'Final score reflects both perception AND reality',
                'Used to classify into risk quadrants'
            ]
        },
        {
            'title': '8. RISK QUADRANT CLASSIFICATION',
            'content': [
                'Elite Equilibrium (H ≥ 70): Green - Excellent Health',
                '  Leadership perception = Reality',
                '  Strong metrics backing strong assessment',
                '  Recommendation: Maintain excellence, scale innovation',
                '',
                'Hidden Excellence (50 ≤ H < 70): Yellow - Good Health',
                '  Leadership underestimating strength',
                '  Good metrics but modest subjective score',
                '  Recommendation: Build confidence, communicate strengths',
                '',
                'Delusional Comfort (30 ≤ H < 50): Orange - At Risk',
                '  Leadership overestimating health',
                '  Perception > Reality',
                '  Recommendation: Address metrics gaps immediately',
                '',
                'Critical Collapse (H < 30): Red - High Risk',
                '  Severe issues in both perception and reality',
                '  Urgent intervention needed',
                '  Recommendation: Comprehensive remedial action plan'
            ]
        }
    ]

    for section in sections:
        ws[f'A{row}'] = section['title']
        apply_cell_style(ws[f'A{row}'], 'header_light', colors)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        for content_line in section['content']:
            ws[f'A{row}'] = content_line
            ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

        row += 1

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

def main():
    print("Creating comprehensive First Opinion Engine Excel workbook...")
    print("  - All 15 challenges with complete questionnaires")
    print("  - Multiple scenario examples with full calculations")
    print("  - Detailed measurement engine explanation")
    print("")

    wb = Workbook()
    wb.remove(wb.active)
    colors = create_colors()

    create_all_challenges_questionnaire(wb, colors)
    print("[OK] Sheet 1: All 15 Challenges Questionnaire")

    create_multiple_scenarios(wb, colors)
    print("[OK] Sheet 2: Scenario Examples & Calculations")

    create_measurement_engine(wb, colors)
    print("[OK] Sheet 3: Measurement Engine Methodology")

    filename = 'public/DISHA_COMPREHENSIVE_FirstOpinion_Engine.xlsx'
    wb.save(filename)

    print("")
    print("=" * 70)
    print("SUCCESS! COMPREHENSIVE FIRST OPINION ENGINE WORKBOOK CREATED")
    print("=" * 70)
    print(f"\nFile saved: {filename}")
    print("\nWorkbook contains:")
    print("  Sheet 1: All 15 Challenges - Complete questionnaires for every challenge")
    print("  Sheet 2: Scenario Examples - 3 different school scenarios with calculations")
    print("  Sheet 3: Measurement Engine - Detailed methodology and formulas")
    print("\nFeatures:")
    print("  - Complete screening questions for all 15 challenges")
    print("  - Multiple real-world scenarios")
    print("  - Step-by-step S_sub calculation")
    print("  - Objective multiplier (M_obj) calculations")
    print("  - Final Health Index with risk classification")
    print("  - Color-coded for easy understanding")
    print("  - All formulas and calculations shown explicitly")
    print("\nReady to open in Microsoft Excel!")

if __name__ == "__main__":
    main()
