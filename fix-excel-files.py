#!/usr/bin/env python3
"""
Fix corrupted Excel files by recreating them with proper structure
"""

def create_excel_files():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        print("Creating First Opinion Engine Excel...")

        # Create First Opinion workbook
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "First Opinion"

        # Title
        ws1['A1'] = "DISHA FIRST OPINION ENGINE"
        ws1['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws1['A1'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        ws1.merge_cells('A1:D1')

        ws1['A2'] = "20-Minute Rapid Institutional Health Assessment"
        ws1.merge_cells('A2:D2')

        # School Details
        ws1['A4'] = "SCHOOL DETAILS"
        ws1['A4'].font = Font(bold=True, color="FFFFFF")
        ws1['A4'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        school_fields = [
            "School Name",
            "Board Affiliation",
            "Total Students",
            "City/District",
            "Annual Fee Tier"
        ]

        for i, field in enumerate(school_fields, 5):
            ws1[f'A{i}'] = field
            ws1[f'A{i}'].font = Font(bold=True)
            ws1[f'B{i}'] = "[Enter value]"
            ws1[f'B{i}'].fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")

        # Challenge Selection
        ws1['A11'] = "CHALLENGE SELECTION (Max 3)"
        ws1['A11'].font = Font(bold=True, color="FFFFFF")
        ws1['A11'].fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")

        challenges = [
            "C1: Enrollment Decline",
            "C2: Student Attrition",
            "C3: Fee Collection",
            "C4: Teacher Attrition",
            "C5: Staff Capability",
            "C6: Leadership Gap",
            "C7: Academic Decline",
            "C8: Student Wellbeing",
            "C9: Remedial Lag",
            "C10: Parent Dissatisfaction",
            "C11: Competitor Pressure",
            "C12: Brand Perception",
            "C13: Cost Inflation",
            "C14: Infra Deficits",
            "C15: Compliance Stress"
        ]

        for i, challenge in enumerate(challenges, 12):
            ws1[f'A{i}'] = challenge
            ws1[f'B{i}'] = "[ ]"

        # Operational Metrics
        ws1['A28'] = "OPERATIONAL METRICS"
        ws1['A28'].font = Font(bold=True, color="FFFFFF")
        ws1['A28'].fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")

        metrics = [
            ("Student-Teacher Ratio", "Ideal: ≤20"),
            ("Parent Response SLA (hours)", "Ideal: ≤12"),
            ("Teacher Training (hours/year)", "Ideal: ≥25"),
            ("Weekly Planning Time (hours)", "Ideal: ≥5")
        ]

        for i, (metric, ideal) in enumerate(metrics, 29):
            ws1[f'A{i}'] = metric
            ws1[f'A{i}'].font = Font(bold=True)
            ws1[f'B{i}'] = "[ENTER]"
            ws1[f'B{i}'].fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
            ws1[f'C{i}'] = ideal

        # Results
        ws1['A34'] = "DIAGNOSTIC RESULTS"
        ws1['A34'].font = Font(bold=True, color="FFFFFF")
        ws1['A34'].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

        results = [
            "Subjective Base Score (S_sub)",
            "Objective Scaling Factor (M_obj)",
            "Delusion Penalty (P_mismatch)",
            "FINAL HEALTH INDEX (H)"
        ]

        for i, result in enumerate(results, 35):
            ws1[f'A{i}'] = result
            ws1[f'A{i}'].font = Font(bold=True)
            ws1[f'B{i}'] = "0-100"
            ws1[f'B{i}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 20

        wb1.save('public/Disha_First_Opinion_Engine.xlsx')
        print("✅ First Opinion Engine created!")

        # Create 14D EWISR workbook
        print("Creating 14D EWISR Master Engine Excel...")

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "14D Assessment"

        ws2['A1'] = "DISHA 14-DIMENSION EWISR ASSESSMENT"
        ws2['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws2['A1'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        ws2.merge_cells('A1:F1')

        ws2['A2'] = "Comprehensive Institutional Health Audit"
        ws2.merge_cells('A2:F2')

        # 14 Dimensions
        dimensions = [
            "D01: Academic Reputation & Rigour",
            "D02: Teacher Welfare & Development",
            "D03: Leadership & Governance Quality",
            "D04: Parent Engagement & SLA",
            "D05: Student Safety & Wellness",
            "D06: Infrastructure & Facilities",
            "D07: Co-Curricular Education",
            "D08: Individual Attention (PTR)",
            "D09: Value for Money",
            "D10: Special Needs Inclusivity",
            "D11: Community Service & Social Responsibility",
            "D12: Faculty Competence & Retention",
            "D13: Internationalism & Cultural Diversity",
            "D14: Management Vision & Growth Drive"
        ]

        ws2['A4'] = "14 DIMENSIONS"
        ws2['A4'].font = Font(bold=True, color="FFFFFF")
        ws2['A4'].fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")

        headers = ["Dimension", "Current", "Benchmark", "Gap", "Priority", "Actions"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for i, dimension in enumerate(dimensions, 6):
            ws2[f'A{i}'] = dimension
            ws2[f'A{i}'].font = Font(bold=True)
            ws2[f'B{i}'] = "[0-100]"
            ws2[f'B{i}'].fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
            ws2[f'C{i}'] = "[BENCHMARK]"
            ws2[f'D{i}'] = "[GAP]"
            ws2[f'E{i}'] = "[H/M/L]"
            ws2[f'F{i}'] = "[ACTIONS]"

        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 25

        wb2.save('public/Disha_14D_EWISR_Master_Engine.xlsx')
        print("✅ 14D EWISR Master Engine created!")

        print("\n" + "="*60)
        print("✅ EXCEL FILES FIXED AND READY!")
        print("="*60)
        print("\n📁 Files created in: public/")
        print("   1. Disha_First_Opinion_Engine.xlsx")
        print("   2. Disha_14D_EWISR_Master_Engine.xlsx")
        print("\n✨ You can now open these in Microsoft Excel!")

    except ImportError as e:
        print(f"❌ Missing dependency: {e}")
        print("\nInstalling openpyxl...")
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        print("✅ Installation complete! Running again...")
        create_excel_files()

if __name__ == "__main__":
    create_excel_files()
